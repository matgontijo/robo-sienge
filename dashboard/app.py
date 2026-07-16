import os
import asyncio
import threading
from datetime import date
from typing import List, Optional
from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks, Response, Request, UploadFile, File, Form
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import config
from dashboard import database as db
from dashboard.auth import get_current_user
import orchestrator

app = FastAPI(title="Robô Conciliação - Dashboard")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Diretório para servir estáticos
static_dir = os.path.join(os.path.dirname(__file__), "static")
if not os.path.exists(static_dir):
    os.makedirs(static_dir)

class RunRequest(BaseModel):
    data_inicio: str
    data_fim: str

from pydantic import Extra
class ConfigUpdate(BaseModel, extra=Extra.allow):
    pass

# ---------------------------------------------------------
# Static HTML e CSS/JS (NÃO EXIGE AUTH HTTP NA ROTA DIRETA 
# POIS O JAVASCRIPT CUIDA DA SESSÃO, EXCETO /api)
# ---------------------------------------------------------
@app.get("/")
async def read_index():
    return FileResponse(os.path.join(static_dir, "index.html"))

app.mount("/static", StaticFiles(directory=static_dir), name="static")

# ---------------------------------------------------------
# Endpoints da API (EXIGEM AUTH)
# ---------------------------------------------------------
api_route = app.router

@app.get("/api/me", dependencies=[Depends(get_current_user)])
def get_me(current_user: db.Usuario = Depends(get_current_user)):
    return {"username": current_user.username, "role": current_user.role}

@app.get("/api/stats", dependencies=[Depends(get_current_user)])
def get_stats():
    return db.get_stats_gerais()

@app.get("/api/execucoes", dependencies=[Depends(get_current_user)])
def listar_execucoes(limit: int = 50):
    return db.get_execucoes(limit)

@app.get("/api/execucoes/{execucao_id}", dependencies=[Depends(get_current_user)])
def obter_execucao(execucao_id: int):
    execucao = db.get_execucao(execucao_id)
    if not execucao:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    
    divergencias = db.get_divergencias(execucao_id)
    return {
        "execucao": execucao,
        "divergencias": divergencias
    }

@app.get("/api/execucoes/{execucao_id}/divergencias", dependencies=[Depends(get_current_user)])
def listar_divergencias(execucao_id: int, criticidade: str = None, q: str = None):
    return db.get_divergencias(execucao_id, criticidade, q)

@app.get("/api/execucoes/{execucao_id}/pagamentos", dependencies=[Depends(get_current_user)])
def listar_pagamentos(execucao_id: int):
    return db.get_pagamentos(execucao_id)

NOMES_TIPO_PT = {
    "PAGAMENTO_DESTINO_DIVERGENTE": "Destino do pagamento diferente do CNPJ do credor",
    "CNPJ_DIVERGENTE": "CNPJ da nota diferente do credor do titulo",
    "VALOR_DIVERGENTE": "Valor do titulo diferente do valor da nota",
    "LIQUIDO_BRUTO_DIVERGENTE": "Liquido diferente de bruto menos retencoes (via NF)",
    "LIQUIDO_PARCELA_DIVERGENTE": "Parcela menos retencoes (caucao/INSS...) diferente do valor a pagar",
    "BOLETO_VALOR_DIVERGENTE": "Valor do boleto (linha digitavel) diferente do valor a pagar",
    "BOLETO_BANCO_INCOMPATIVEL": "Banco do boleto incompativel com a forma (proprio/outros)",
    "BOLETO_VENCIMENTO_DIVERGENTE": "Vencimento do boleto diferente do titulo",
    "TRANSFERENCIA_BANCO_INCOMPATIVEL": "TED x deposito mesmo banco (conta destino)",
    "IMPOSTO_DIVERGENTE": "Imposto retido diferente do destacado na nota",
    "CHAVE_NFE_INVALIDA": "Chave de NF-e invalida",
    "BOLETO_NAO_ENCONTRADO": "Boleto nao encontrado no DDA",
    "SEM_ANEXO": "Titulo sem anexo no Sienge",
    "ANEXO_ILEGIVEL": "Anexo nao lido (OCR pendente)",
    "PIX_NAO_VERIFICAVEL": "Chave Pix nao verificavel",
    "PAGAMENTO_FORMA_INCOMPATIVEL": "Forma de pagamento incompativel",
    "FORMA_PAGAMENTO_AUSENTE": "Sem forma de pagamento cadastrada",
    "NF_SEM_CNPJ_DO_CREDOR": "CNPJ do credor nao aparece na NF anexada",
    "IMPOSTO_NF_DIVERGENTE": "Retencao do titulo diferente da destacada na nota",
    "IMPOSTO_NAO_RETIDO": "Nota destaca retencao que o titulo nao lancou",
    "RETENCAO_ALIQUOTA_SUSPEITA": "Aliquota de retencao acima do usual",
    "RETENCAO_INDEVIDA_SIMPLES": "Retencao federal de fornecedor do Simples",
    "VENCIMENTO_DIVERGENTE": "Vencimento divergente",
}

@app.get("/api/execucoes/{execucao_id}/revisao/relatorio", dependencies=[Depends(get_current_user)])
def relatorio_revisao(execucao_id: int, status: str = "REJEITADO"):
    """Gera um Excel com as divergências revisadas no status pedido
    (REJEITADO por padrão) — o dossiê para corrigir os lançamentos no Sienge."""
    status = (status or "REJEITADO").upper()
    divs = [d for d in db.get_divergencias(execucao_id)
            if (d.status_revisao or "PENDENTE") == status]

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    import datetime as _dt

    wb = Workbook()
    ws = wb.active
    ws.title = status.capitalize() + "s"
    headers = ["Nº Título", "Fornecedor", "CNPJ Fornecedor", "Valor (R$)", "Vencimento",
               "Problema", "Campo", "Sienge diz", "Verificado", "Criticidade",
               "Observação da revisão", "Revisado por", "Revisado em"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="D9D9D9")
    for d in divs:
        ws.append([
            d.titulo_numero, d.fornecedor_nome, d.fornecedor_cnpj,
            d.valor_sienge, str(d.data_vencimento or ""),
            NOMES_TIPO_PT.get(d.tipo, d.tipo), d.campo,
            d.valor_sienge_campo, d.valor_boleto_campo if d.valor_boleto_campo not in (None, "-") else d.valor_nfe_campo,
            d.criticidade, d.observacao_revisao,
            d.revisado_por, str(d.revisado_em or "")[:16].replace("T", " "),
        ])
    ws.auto_filter.ref = f"A1:M{max(1, len(divs) + 1)}"
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(width + 2, 55)

    pasta = os.path.join(config.OUTPUT_DIR, "relatorios")
    os.makedirs(pasta, exist_ok=True)
    nome = f"revisao_{status.lower()}_ciclo{execucao_id}_{_dt.datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
    caminho = os.path.join(pasta, nome)
    wb.save(caminho)
    return FileResponse(path=caminho, filename=nome,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

class RevisaoPayload(BaseModel):
    status: str  # PENDENTE | APROVADO | REJEITADO
    observacao: Optional[str] = None

@app.post("/api/divergencias/{divergencia_id}/revisao")
def revisar_divergencia(divergencia_id: int, payload: RevisaoPayload,
                        user=Depends(get_current_user)):
    status_norm = (payload.status or "").upper()
    if status_norm not in ("PENDENTE", "APROVADO", "REJEITADO"):
        raise HTTPException(status_code=400, detail="status deve ser PENDENTE, APROVADO ou REJEITADO")
    d = db.atualizar_revisao(divergencia_id, status_norm, payload.observacao,
                             getattr(user, "username", None))
    if not d:
        raise HTTPException(status_code=404, detail="Divergência não encontrada")
    return d

@app.get("/api/execucoes/{execucao_id}/logs", dependencies=[Depends(get_current_user)])
def listar_logs(execucao_id: int):
    return db.get_logs(execucao_id)

@app.get("/api/execucoes/{execucao_id}/relatorio", dependencies=[Depends(get_current_user)])
def download_relatorio(execucao_id: int):
    execucao = db.get_execucao(execucao_id)
    if not execucao or not execucao.relatorio_path:
        raise HTTPException(status_code=404, detail="Relatório não encontrado")
        
    path = execucao.relatorio_path
    
    # Path traversal protection
    abs_path = os.path.abspath(path)
    abs_output = os.path.abspath(config.OUTPUT_DIR)
    if not abs_path.startswith(abs_output):
        raise HTTPException(status_code=403, detail="Acesso negado")
        
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="Arquivo físico não encontrado")
        
    return FileResponse(
        path=abs_path, 
        filename=os.path.basename(abs_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.get("/api/execucoes/{execucao_id}/abrir-pasta", dependencies=[Depends(get_current_user)])
def abrir_pasta_anexos(execucao_id: int):
    """Abre, no Explorer da máquina que roda o painel (o PC do usuário), a pasta
    com os documentos deste ciclo: o dossiê (NF/boleto organizados) se existir,
    senão a pasta geral de anexos. Retorna o caminho para exibir/copiar."""
    import sys
    dossie = os.path.abspath(os.path.join(config.OUTPUT_DIR, "dossie", f"ciclo_{execucao_id}"))
    anexos = os.path.abspath(os.path.join(config.OUTPUT_DIR, "anexos"))
    pasta = dossie if os.path.isdir(dossie) else anexos
    existe = os.path.isdir(pasta)
    aberto = False
    if existe and sys.platform.startswith("win") and hasattr(os, "startfile"):
        try:
            os.startfile(pasta)  # noqa: S606 — ferramenta local, abre o Explorer
            aberto = True
        except Exception:  # noqa: BLE001
            aberto = False
    return {"pasta": pasta, "existe": existe, "aberto": aberto}

@app.get("/api/execucoes/{execucao_id}/danfe", dependencies=[Depends(get_current_user)])
def download_danfe(execucao_id: int, path: str):
    # Path traversal protection
    abs_path = os.path.abspath(path)
    abs_output = os.path.abspath(config.OUTPUT_DIR)
    if not abs_path.startswith(abs_output):
        raise HTTPException(status_code=403, detail="Acesso negado")
        
    if not os.path.exists(abs_path):
        raise HTTPException(status_code=404, detail="Arquivo físico não encontrado")
        
    return FileResponse(path=abs_path, filename=os.path.basename(abs_path), media_type="application/pdf")

def _limpar_execucoes_orfas():
    """Execução 'RODANDO' sem log novo há 15+ min é zumbi (processo caiu/foi
    morto) — marca como ABORTADO para não travar novas execuções com 409."""
    import datetime as _dt
    agora = _dt.datetime.now()
    for e in db.get_execucoes(limit=20):
        exec_id = getattr(e, "id", None)
        if e.status != "RODANDO" or exec_id is None:
            continue
        logs = db.get_logs(exec_id)
        ultimo = logs[-1].timestamp if logs else getattr(e, "iniciado_em", None)
        if ultimo and (agora - ultimo).total_seconds() > 900:
            db.atualizar_execucao(exec_id, status="ABORTADO", concluido_em=agora,
                                  erro_mensagem="Processo interrompido — marcado automaticamente como abortado")

_limpar_execucoes_orfas()  # limpeza na subida do painel

@app.post("/api/execucoes/iniciar", dependencies=[Depends(get_current_user)])
def iniciar_execucao(req: RunRequest, current_user: db.Usuario = Depends(get_current_user)):
    if current_user.role == "LEITURA":
        raise HTTPException(status_code=403, detail="Usuário sem permissão para iniciar execução")

    # Verifica se já tem alguma execução rodando (limpando zumbis antes)
    _limpar_execucoes_orfas()
    execs = db.get_execucoes(limit=10)
    for e in execs:
        if e.status == "RODANDO":
            raise HTTPException(status_code=409, detail="Já existe uma execução em andamento")
            
    d_inicio = date.fromisoformat(req.data_inicio)
    d_fim = date.fromisoformat(req.data_fim)
    
    # Inicia a execução do orchestrator em uma thread separada para não bloquear a API
    def rotina_em_background():
        orchestrator.executar_ciclo(d_inicio, d_fim, iniciado_por="dashboard")
        
    thread = threading.Thread(target=rotina_em_background)
    thread.start()
    
    # Precisamos retornar algo para o frontend, mas o ID só é gerado DENTRO do orquestrador.
    # O ideal era criar a execução aqui e passar o ID pro orquestrador, mas para simplificar
    # vamos aguardar uma fração de segundo até que a thread crie o ID.
    import time
    time.sleep(0.5)
    
    # Pega a ultima (que deve ser a nossa)
    ultima = db.get_execucoes(limit=1)
    if ultima and ultima[0].status == "RODANDO":
        return {"execucao_id": ultima[0].id}
    return {"execucao_id": 0}

@app.post("/api/execucoes/iniciar-relatorio", dependencies=[Depends(get_current_user)])
def iniciar_execucao_relatorio(
    arquivo: UploadFile = File(...),
    data_inicio: str = Form(None),
    data_fim: str = Form(None),
    current_user: db.Usuario = Depends(get_current_user),
):
    """Inicia uma execução usando um relatório de Fluxo de Caixa enviado por upload."""
    if current_user.role == "LEITURA":
        raise HTTPException(status_code=403, detail="Usuário sem permissão para iniciar execução")

    ext = os.path.splitext(arquivo.filename or "")[1].lower()
    if ext not in (".xlsx", ".xlsm", ".pdf"):
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um .xlsx ou .pdf")

    # Já existe execução rodando? (limpando zumbis antes)
    _limpar_execucoes_orfas()
    for e in db.get_execucoes(limit=10):
        if e.status == "RODANDO":
            raise HTTPException(status_code=409, detail="Já existe uma execução em andamento")

    # Salva o relatório enviado
    import time as _time
    upload_dir = os.path.join(config.OUTPUT_DIR, "relatorios_input")
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = f"{_time.strftime('%Y%m%d_%H%M%S')}{ext}"
    dest = os.path.join(upload_dir, safe_name)
    with open(dest, "wb") as f:
        f.write(arquivo.file.read())

    # Datas: usadas apenas para a janela DDA; default hoje -> +7 dias
    from datetime import timedelta
    try:
        d_inicio = date.fromisoformat(data_inicio) if data_inicio else date.today()
        d_fim = date.fromisoformat(data_fim) if data_fim else d_inicio + timedelta(days=7)
    except ValueError:
        raise HTTPException(status_code=400, detail="Datas inválidas — use o formato YYYY-MM-DD ou deixe em branco")

    def rotina_em_background():
        orchestrator.executar_ciclo(d_inicio, d_fim, iniciado_por="dashboard-relatorio", relatorio_path=dest)

    threading.Thread(target=rotina_em_background).start()

    import time
    time.sleep(0.5)
    ultima = db.get_execucoes(limit=1)
    if ultima and ultima[0].status == "RODANDO":
        return {"execucao_id": ultima[0].id}
    return {"execucao_id": 0}

@app.post("/api/execucoes/{execucao_id}/abortar", dependencies=[Depends(get_current_user)])
def abortar_execucao(execucao_id: int, current_user: db.Usuario = Depends(get_current_user)):
    if current_user.role == "LEITURA":
        raise HTTPException(status_code=403, detail="Usuário sem permissão para abortar execução")
        
    execucao = db.get_execucao(execucao_id)
    if not execucao:
        raise HTTPException(status_code=404, detail="Execução não encontrada")
    if execucao.status != "RODANDO":
        raise HTTPException(status_code=400, detail="Execução não está rodando")
        
    orchestrator.abortar_execucao(execucao_id)
    return {"status": "Abort signal sent"}

# ---------------------------------------------------------
# Configurações do .env
# ---------------------------------------------------------
import dotenv

@app.get("/api/config", dependencies=[Depends(get_current_user)])
def get_config(current_user: db.Usuario = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Apenas ADMIN pode visualizar as configurações")
        
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    if not os.path.exists(env_path):
        return {}
    return dotenv.dotenv_values(env_path)

@app.post("/api/config", dependencies=[Depends(get_current_user)])
def update_config(payload: dict, current_user: db.Usuario = Depends(get_current_user)):
    if current_user.role != "ADMIN":
        raise HTTPException(status_code=403, detail="Apenas ADMIN pode alterar as configurações")
        
    env_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), ".env")
    if not os.path.exists(env_path):
        open(env_path, 'a').close()
        
    for k, v in payload.items():
        dotenv.set_key(env_path, k, str(v))
        
    config.reload_config()
    return {"status": "success", "message": "Configurações salvas e recarregadas."}

# ---------------------------------------------------------
# SSE - Server Sent Events (TEMPO REAL)
# O sse_starlette.sse lida com isso elegantemente
# ---------------------------------------------------------
@app.get("/api/stream/{execucao_id}")
async def stream_logs(execucao_id: int, req: Request):
    
    # Valida Auth no cabeçalho ou cookie para o SSE (geralmente enviamos o header de auth)
    # A dependência Depends() falha no SSE nativo do navegador as vezes, por isso
    # o frontend tem que enviar no EventSource se possível ou usar cookies.
    # Com a autenticação desligada (uso local), o stream é aberto direto.
    if config.DASHBOARD_AUTH:
        auth = req.headers.get("Authorization")
        if not auth:
            auth = req.query_params.get("token")
            if auth and auth not in ("null", "undefined"):
                auth = f"Basic {auth}"
            else:
                auth = None

        if not auth:
            raise HTTPException(status_code=401, detail="Login necessário")

        import base64
        try:
            scheme, credentials = auth.split()
            if scheme.lower() != 'basic':
                raise Exception()
            decoded = base64.b64decode(credentials).decode("ascii")
            username, _, password = decoded.partition(":")

            db_session = db.SessionLocal()
            try:
                user = db_session.query(db.Usuario).filter(db.Usuario.username == username).first()
                if not user or not db.pwd_context.verify(password, user.password_hash):
                    raise Exception()
            finally:
                db_session.close()

        except Exception:
            # NUNCA envia WWW-Authenticate: é isso que dispara o pop-up nativo
            # de login do navegador por cima do painel.
            raise HTTPException(status_code=401, detail="Login inválido")

    async def event_generator():
        last_id = 0
        while True:
            # Checa conexao
            if await req.is_disconnected():
                break
                
            # Buscar novos logs
            novos_logs = db.get_logs(execucao_id, last_id=last_id)
            for l in novos_logs:
                data_str = f"[{l.timestamp.strftime('%H:%M:%S')}] {l.level.ljust(8)} | {l.modulo} - {l.mensagem}"
                yield f"event: log\ndata: {data_str}\n\n"
                last_id = l.id
                
            # Verifica se terminou
            e = db.get_execucao(execucao_id)
            if not e or e.status != "RODANDO":
                yield f"event: close\ndata: Fechando stream\n\n"
                break
                
            await asyncio.sleep(1)

    return StreamingResponse(event_generator(), media_type="text/event-stream")
