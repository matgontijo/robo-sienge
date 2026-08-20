"""Gera o PDF do Fluxo de Caixa Analitico no padrao Sienge, a partir da planilha
do ciclo (CICLO1_SEPARADOS.xlsx / CICLO2_REMESSA.xlsx).

Layout medido no relatorio original (relatorio (19).pdf):
  A4 retrato; Helvetica 7 no corpo, 16 no titulo, 6 no rodape;
  grade com traco 0.25 preto; colunas em
  36 | 77.8 | 156.3 | 213.8 | 234.7 | 386.4 | 443.9 | 501.5 | 559;
  linha de 9.1pt (cresce em multiplos quando o nome do fornecedor quebra);
  blocos cinza (0.749) nos rotulos do cabecalho;
  cabecalho da tabela repetido apos cada 'Total do dia';
  rodape 'data | SIENGE / STARIAN | n de total'.

Uso:
    python gerar_pdf_fluxo.py output/CICLO1_SEPARADOS.xlsx
    python gerar_pdf_fluxo.py output/CICLO2_REMESSA.xlsx saida.pdf
"""
import os
import re
import sys
from datetime import datetime

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib.utils import ImageReader
from reportlab.pdfgen import canvas

LARG, ALT = A4
CINZA = 0.749
FS, FS_TIT, FS_ROD = 7, 16, 6
TRACO = 0.25
H = 9.1                                   # altura da linha
Y_TAB = 188.7                             # base do cabecalho da tabela
Y_RODAPE = 796.5
Y_LIM = 780.0                             # ultima base utilizavel antes do rodape

# limites das colunas da tabela
X = [36.0, 77.8, 156.3, 213.8, 234.7, 386.4, 443.9, 501.5, 559.0]
# limites do bloco de cabecalho (4 colunas)
XC = [36.0, 166.8, 297.5, 428.2, 559.0]
PAD = 1.3
LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "assets", "logo_relatorio.png")


def dinheiro(v):
    return f"{float(v or 0):,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")


class PDF:
    def __init__(self, caminho, cab, total_paginas=None):
        self.c = canvas.Canvas(caminho, pagesize=A4)
        self.c.setLineWidth(TRACO)
        self.cab = cab
        self.pag = 0
        self.y = 0
        self.total_paginas = total_paginas

    # ---------------- primitivas ----------------
    def txt(self, x, y, s, bold=False, size=FS, dir=False):
        self.c.setFillGray(0)
        self.c.setFont("Helvetica-Bold" if bold else "Helvetica", size)
        (self.c.drawRightString if dir else self.c.drawString)(x, ALT - y, str(s))

    def cel(self, x0, x1, y, alt=H, cinza=False, borda=True):
        """Celula da grade: y e a BASE do texto; a caixa vai de y-alt+2 a y+2."""
        top, bot = y - alt + 2.0, y + 2.0
        if cinza:
            self.c.setFillGray(CINZA)
            self.c.rect(x0, ALT - bot, x1 - x0, bot - top, stroke=0, fill=1)
        if borda:
            self.c.setStrokeGray(0)
            self.c.rect(x0, ALT - bot, x1 - x0, bot - top, stroke=1, fill=0)
        self.c.setFillGray(0)

    # ---------------- blocos ----------------
    def nova_pagina(self):
        if self.pag:
            self.rodape()
            self.c.showPage()
            self.c.setLineWidth(TRACO)
        self.pag += 1
        self.topo()
        self.cab_tabela(Y_TAB)
        self.y = Y_TAB + H

    def rodape(self):
        self.txt(37, Y_RODAPE, self.cab["emissao"], size=FS_ROD)
        self.txt(271.7, Y_RODAPE, "SIENGE / STARIAN", size=FS_ROD)
        total = self.total_paginas or self.pag
        self.txt(534.6, Y_RODAPE, f"{self.pag} de {total}", size=FS_ROD)

    def topo(self):
        # caixa do logo + caixa do titulo
        self.c.setStrokeGray(0)
        self.c.rect(36, ALT - 95, 114, 58, stroke=1, fill=0)
        self.c.rect(150, ALT - 95, 409, 58, stroke=1, fill=0)
        if os.path.exists(LOGO):
            try:
                self.c.drawImage(ImageReader(LOGO), 39.05, ALT - 91.08,
                                 width=110.02, height=54.08, mask="auto")
            except Exception:
                pass
        larg = self.c.stringWidth(self.cab["titulo"], "Helvetica-Bold", FS_TIT)
        self.txt(150 + (409 - larg) / 2, 72, self.cab["titulo"], bold=True, size=FS_TIT)

        # bloco de identificacao (4 colunas)
        linhas = [("Fluxo de caixa por", self.cab["por"], "Período", self.cab["periodo"], 1),
                  ("Valores corrigidos em", "REAL", "Valores apresentados em", "REAL", 1),
                  ("Grupo de empresa", self.cab["grupo"], "Empresa", self.cab["empresa"], 3),
                  ("Tipo de análise", self.cab["tipo"], None, None, 1)]
        y = 111.9
        for rot1, v1, rot2, v2, n in linhas:
            alt = H * n
            self.cel(XC[0], XC[1], y + alt - H, alt, cinza=True)
            self.cel(XC[1], XC[2], y + alt - H, alt)
            self.txt(XC[1] - PAD, y, rot1, bold=True, dir=True)
            self.txt(XC[1] + PAD, y, v1)
            if rot2:
                self.cel(XC[2], XC[3], y + alt - H, alt, cinza=True)
                self.cel(XC[3], XC[4], y + alt - H, alt)
                self.txt(XC[3] - PAD, y, rot2, bold=True, dir=True)
                for k, parte in enumerate(quebrar(self.c, v2, XC[4] - XC[3] - 2 * PAD)):
                    self.txt(XC[3] + PAD, y + k * H, parte)
            y += alt

        # disponivel
        yd = y + H
        self.cel(X[0], 501.5, yd, cinza=True)
        self.cel(501.5, X[-1], yd)
        self.txt(X[0] + PAD, yd, self.cab["disponivel_rot"], bold=True)
        self.txt(X[-1] - PAD, yd, self.cab["disponivel_val"], dir=True)

    def cab_tabela(self, y):
        rotulos = [("Data", False), ("Documento", False), ("Tit/Parc", False),
                   ("Orig.", False), ("Cliente/Fornecedor", False),
                   ("Entradas", True), ("Saídas", True), ("Saldo", True)]
        for i, (rot, dir) in enumerate(rotulos):
            self.cel(X[i], X[i + 1], y)
            self.txt(X[i + 1] - PAD if dir else X[i] + PAD, y, rot, bold=True, dir=dir)

    def linha(self, r):
        partes = quebrar(self.c, r["forn"], X[5] - X[4] - 2 * PAD)
        alt = H * len(partes)
        if self.y + alt > Y_LIM:
            self.nova_pagina()
        for i in range(8):
            self.cel(X[i], X[i + 1], self.y + alt - H, alt)
        self.txt(X[0] + PAD, self.y, r["data"])
        self.txt(X[1] + PAD, self.y, r["doc"])
        self.txt(X[2] + PAD, self.y, r["tp"])
        self.txt(X[3] + PAD, self.y, r["orig"])
        for k, p in enumerate(partes):
            self.txt(X[4] + PAD, self.y + k * H, p)
        self.txt(X[6] - PAD, self.y, dinheiro(r["ent"]), dir=True)
        self.txt(X[7] - PAD, self.y, dinheiro(r["sai"]), dir=True)
        self.txt(X[8] - PAD, self.y, dinheiro(r["saldo"]), dir=True)
        self.y += alt

    def total(self, rotulo, ent, sai, repetir_cab=True):
        if self.y + H * 3 > Y_LIM:
            self.nova_pagina()
        self.cel(X[0], X[5], self.y, cinza=True)
        self.cel(X[5], X[6], self.y, cinza=True)
        self.cel(X[6], X[7], self.y, cinza=True)
        self.txt(X[0] + PAD, self.y, rotulo, bold=True)
        self.txt(X[6] - PAD, self.y, dinheiro(ent), bold=True, dir=True)
        self.txt(X[7] - PAD, self.y, dinheiro(sai), bold=True, dir=True)
        self.y += H * 2
        if repetir_cab:
            self.cab_tabela(self.y)
            self.y += H

    def salvar(self):
        self.rodape()
        self.c.save()


def quebrar(c, texto, larg, fonte="Helvetica", size=FS):
    palavras = str(texto or "").split()
    linhas, atual = [], ""
    for p in palavras:
        teste = (atual + " " + p).strip()
        if c.stringWidth(teste, fonte, size) <= larg or not atual:
            atual = teste
        else:
            linhas.append(atual)
            atual = p
    linhas.append(atual)
    return linhas or [""]


def ler_planilha(caminho):
    ws = load_workbook(caminho, data_only=True).active
    NCOL = ws.max_column
    lin_cab = next(i for i in range(1, ws.max_row + 1)
                   if any(ws.cell(row=i, column=c).value and
                          "Tit/Parc" in str(ws.cell(row=i, column=c).value)
                          for c in range(1, NCOL + 1)))
    cel = lambda l, c: str(ws.cell(row=l, column=c).value or "")
    cab = {"titulo": cel(2, 4) or "Fluxo de Caixa Analítico",
           "por": cel(5, 5), "periodo": cel(5, 15),
           "grupo": cel(7, 5), "empresa": cel(7, 15),
           "tipo": cel(8, 5) or "A realizar",
           "disponivel_rot": cel(10, 1),
           "disponivel_val": dinheiro(ws.cell(row=10, column=18).value)}
    itens = []
    for i in range(lin_cab + 1, ws.max_row + 1):
        tp = cel(i, 4).strip()
        rot = cel(i, 1).strip()
        if re.fullmatch(r"\d+/\d+", tp):
            itens.append({"tipo": "linha", "data": cel(i, 1)[:10], "doc": cel(i, 2),
                          "tp": tp, "orig": cel(i, 7), "forn": cel(i, 8),
                          "ent": ws.cell(row=i, column=13).value or 0,
                          "sai": ws.cell(row=i, column=16).value or 0,
                          "saldo": ws.cell(row=i, column=18).value or 0})
        elif rot.lower().startswith("total"):
            itens.append({"tipo": "total", "rotulo": rot,
                          "ent": ws.cell(row=i, column=13).value or 0,
                          "sai": ws.cell(row=i, column=16).value or 0,
                          "periodo": rot.lower().startswith("total do per")})
    return cab, itens


def _render(saida, cab, itens, total_paginas=None):
    pdf = PDF(saida, cab, total_paginas)
    pdf.nova_pagina()
    for it in itens:
        if it["tipo"] == "linha":
            pdf.linha(it)
        else:
            pdf.total(it["rotulo"], it["ent"], it["sai"], repetir_cab=not it["periodo"])
    pdf.salvar()
    return pdf.pag


def gerar(entrada, saida=None):
    """Renderiza duas vezes: a 1a so para contar paginas (o rodape traz 'n de total')."""
    saida = saida or os.path.splitext(entrada)[0] + ".pdf"
    cab, itens = ler_planilha(entrada)
    cab["emissao"] = datetime.now().strftime("%d/%m/%Y - %H:%M:%S")
    total = _render(saida, cab, itens)
    _render(saida, cab, itens, total)
    return saida, total, len([i for i in itens if i["tipo"] == "linha"])


if __name__ == "__main__":
    ent = sys.argv[1] if len(sys.argv) > 1 else "output/CICLO1_SEPARADOS.xlsx"
    caminho, pags, linhas = gerar(ent, sys.argv[2] if len(sys.argv) > 2 else None)
    print(f"gerado: {caminho} | {pags} pagina(s) | {linhas} lancamentos")
