# -*- coding: utf-8 -*-
"""Planilha 'Pagamentos por fornecedor' — uma aba por empresa (Consplan, Cilicon, MVCG,
Alves Filho) + RESUMO. Fonte: output/_fornecedores_4.json e output/_destinos_ted_4.json
(API Sienge). Totais por fórmula; linhas compactas, zebra e autofiltro."""
import json
from datetime import datetime

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

D = json.load(open("output/_fornecedores_4.json", encoding="utf-8"))
TED = json.load(open("output/_destinos_ted_4.json", encoding="utf-8"))

VERDE = "1E6B4F"
F_TIT = Font(name="Arial", bold=True, size=13, color="FFFFFF")
F_TIT2 = Font(name="Arial", size=9, color="D8EAE2")
F_SUB = Font(name="Arial", size=8.5, color="5E6F7C", italic=True)
F_H = Font(name="Arial", bold=True, size=8.5, color="FFFFFF")
F_C = Font(name="Arial", size=9)
F_M = Font(name="Arial", size=9, color="5E6F7C")
F_B = Font(name="Arial", bold=True, size=9)
F_OK = Font(name="Arial", size=9, color="1E6B4F")
F_ABERTO = Font(name="Arial", size=9, color="B0442B", bold=True)
FILL_TIT = PatternFill("solid", fgColor=VERDE)
FILL_H = PatternFill("solid", fgColor="16211B")
FILL_Z = PatternFill("solid", fgColor="F2F5F3")
FILL_TOT = PatternFill("solid", fgColor="DFF1E5")
FILL_PCT = PatternFill("solid", fgColor="FBF0D3")
LATERAL = Side(style="thin", color="E3E8E4")
BD = Border(bottom=LATERAL)
MOEDA = 'R$ #,##0.00;[Red](R$ #,##0.00);"-"'
DATA = "DD/MM/YYYY"
HOJE = datetime.now().strftime("%d/%m/%Y")
DIR = Alignment(horizontal="right")


def corta(s, n):
    s = (s or "").strip()
    return s if len(s) <= n else s[: n - 1] + "…"


def faixa_titulo(ws, nome, sub, ncols):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws["A1"] = nome; ws["A1"].font = F_TIT; ws["A1"].alignment = Alignment(vertical="center")
    ws["A2"] = sub; ws["A2"].font = F_TIT2
    for r in (1, 2):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).fill = FILL_TIT
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 14
    ws.sheet_view.showGridLines = False


def cabecalho(ws, linha, cols, largs, filtro=True):
    for i, (c, w) in enumerate(zip(cols, largs), 1):
        cel = ws.cell(row=linha, column=i, value=c)
        cel.font = F_H; cel.fill = FILL_H
        cel.alignment = Alignment(vertical="center", horizontal=("right" if "R$" in c else "left"))
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[linha].height = 18
    ws.freeze_panes = ws.cell(row=linha + 1, column=1)
    if filtro:
        ws.auto_filter.ref = f"A{linha}:{get_column_letter(len(cols))}{linha}"


def celula(ws, r, c, valor, fonte=None, fmt=None, alinh=None, dica=None):
    cel = ws.cell(row=r, column=c, value=valor)
    cel.font = fonte or F_C
    cel.border = BD
    if fmt:
        cel.number_format = fmt
    if alinh:
        cel.alignment = alinh
    if dica:
        cel.comment = Comment(dica, "Robô Sienge")
    return cel


def data_br(ws, r, c, iso, fonte=None):
    if iso:
        return celula(ws, r, c, datetime.strptime(iso[:10], "%Y-%m-%d"), fonte=fonte, fmt=DATA)
    return celula(ws, r, c, None, fonte=fonte)


def fonte_sit(sit):
    return F_OK if sit == "Totalmente paga" else (F_ABERTO if sit == "Não paga" else F_C)


wb = Workbook()
ORDEM = ["CONSPLAN", "CILICON", "MVCG", "ALVES FILHO"]
refs = {}

COLS = ["Título", "Tipo", "Documento", "Emissão", "Parc.", "Vencimento", "Valor (R$)", "Pago (R$)",
        "Situação", "Forma", "Planilha de custo (obra)", "CC · Plano financeiro", "Lançado por"]
LARGS = [8, 6, 20, 10.5, 5, 10.5, 13, 13, 14, 15, 38, 36, 16]

for apelido in ORDEM:
    v = D[apelido]
    aba = apelido.title() if apelido != "MVCG" else "MVCG"
    ws = wb.active if wb.sheetnames == ["Sheet"] else wb.create_sheet(aba)
    ws.title = aba

    if apelido == "MVCG":
        cols = ["Fornecedor do título", "Título", "Doc", "Parc.", "Vencimento", "Valor (R$)", "Pago (R$)", "Situação", "Conta favorecida"]
        largs = [22, 8, 16, 5, 10.5, 13, 13, 14, 26]
        faixa_titulo(ws, v["nome"], f"CNPJ {v['cnpj']} · representante de pagamento da Consplan · fonte: API Sienge · {HOJE}", len(cols))
        ws["A3"] = ("Sem título próprio no contas a pagar: a MVCG recebe as TEDs dos títulos da CONSPLAN "
                    "(Itaú 341 · ag 0705 · c/c 9091-8) — arranjo cadastrado, não é desvio.")
        ws["A3"].font = F_SUB
        cabecalho(ws, 5, cols, largs)
        r = 6
        mv = sorted([d for d in TED if "MVCG" in (d.get("favorecido") or "").upper()],
                    key=lambda d: (d["venc"] or "", d["titulo"]))
        for k, d in enumerate(mv):
            zebra = FILL_Z if k % 2 else None
            celula(ws, r, 1, "CONSPLAN")
            celula(ws, r, 2, d["titulo"])
            celula(ws, r, 3, corta(d["doc"], 15), fonte=F_M)
            celula(ws, r, 4, d["parcela"])
            data_br(ws, r, 5, d["venc"])
            celula(ws, r, 6, d["valor"], fmt=MOEDA)
            celula(ws, r, 7, f'=IF($H{r}="Totalmente paga",$F{r},0)', fmt=MOEDA)
            celula(ws, r, 8, d["sit"], fonte=fonte_sit(d["sit"]))
            celula(ws, r, 9, f"Itaú ag {d['agencia']} c/c {d['conta']}", fonte=F_M)
            if zebra:
                for c in range(1, len(cols) + 1):
                    ws.cell(row=r, column=c).fill = zebra
            ws.row_dimensions[r].height = 14.5
            r += 1
        fim = r - 1
        celula(ws, r, 5, "TOTAL", fonte=F_B)
        celula(ws, r, 6, f"=SUM(F6:F{fim})", fonte=F_B, fmt=MOEDA)
        celula(ws, r, 7, f"=SUM(G6:G{fim})", fonte=F_B, fmt=MOEDA)
        for c in range(1, len(cols) + 1):
            ws.cell(row=r, column=c).fill = FILL_TOT
        ws.row_dimensions[r].height = 16
        refs[apelido] = dict(lanc=f"'{aba}'!F{r}", pago=f"'{aba}'!G{r}", aberto=None, n=len(mv))
        continue

    faixa_titulo(ws, v["nome"], f"CNPJ {v['cnpj']} · credor {v['credor_id']} · histórico 01/01/2019 a {HOJE} · fonte: API Sienge", len(COLS))
    cabecalho(ws, 4, COLS, LARGS)
    r = 5
    k = 0
    for t in v["titulos"]:
        for p in sorted(t["parcelas"], key=lambda x: (x["venc"] or "", x["n"])):
            eh_pct = t["tipo"] == "PCT"
            zebra = FILL_PCT if eh_pct else (FILL_Z if k % 2 else None)
            celula(ws, r, 1, t["id"])
            celula(ws, r, 2, t["tipo"], fonte=(F_ABERTO if eh_pct else F_M))
            celula(ws, r, 3, corta(t["doc"], 18), fonte=F_M, dica=(t["doc"] if len(t["doc"] or "") > 18 else None))
            data_br(ws, r, 4, t["emissao"], fonte=F_M)
            celula(ws, r, 5, p["n"])
            data_br(ws, r, 6, p["venc"])
            celula(ws, r, 7, p["valor"], fmt=MOEDA)
            celula(ws, r, 8, f'=IF($I{r}="Totalmente paga",$G{r},0)', fmt=MOEDA)
            celula(ws, r, 9, p["sit"], fonte=fonte_sit(p["sit"]))
            celula(ws, r, 10, corta((p["forma"] or "").replace("Boleto Bancário", "Boleto"), 20), fonte=F_M)
            celula(ws, r, 11, corta(t["planilha"], 40), fonte=F_M,
                   dica=(f"{t['obra']}\n{t['planilha']}" if (len(t["planilha"] or "") > 40) else None))
            celula(ws, r, 12, corta(t["cc_plano"], 40), fonte=F_M, dica=(t["cc_plano"] if len(t["cc_plano"] or "") > 40 else None))
            celula(ws, r, 13, corta((t["lancado_por"] or "").title(), 18), fonte=F_M)
            if zebra:
                for c in range(1, len(COLS) + 1):
                    ws.cell(row=r, column=c).fill = zebra
            ws.row_dimensions[r].height = 14.5
            r += 1
        k += 1
    fim = r - 1
    totais = [
        ("TOTAL LANÇADO (todas as parcelas)", 7, f"=SUM(G5:G{fim})"),
        ("TOTAL PAGO", 8, f"=SUM(H5:H{fim})"),
        ("PAGO SEM PCT (sem duplicar o contrato)", 8, f'=SUMIFS(H5:H{fim},B5:B{fim},"<>PCT")'),
        ("EM ABERTO SEM PCT", 7, f'=SUMIFS(G5:G{fim},B5:B{fim},"<>PCT")-SUMIFS(H5:H{fim},B5:B{fim},"<>PCT")'),
    ]
    for rot, col, formula in totais:
        celula(ws, r, 4, rot, fonte=F_B)
        celula(ws, r, col, formula, fonte=F_B, fmt=MOEDA)
        for c in range(1, len(COLS) + 1):
            ws.cell(row=r, column=c).fill = FILL_TOT
        ws.row_dimensions[r].height = 16
        r += 1
    celula(ws, r + 1, 4, "Linhas amarelas = PCT (contrato): as baixas repetem os pagamentos das medições — use os totais 'sem PCT'.", fonte=F_SUB)
    refs[apelido] = dict(lanc=f"'{aba}'!G{fim+1}", pago_sem=f"'{aba}'!H{fim+3}", aberto=f"'{aba}'!G{fim+4}", n=len(v["titulos"]))

# ============================== RESUMO ==============================
ws = wb.create_sheet("RESUMO", 0)
COLS_R = ["Empresa", "CNPJ", "Títulos", "Total lançado (R$)", "Pago sem PCT (R$)", "Em aberto sem PCT (R$)"]
faixa_titulo(ws, "Pagamentos por fornecedor — Grupo Garden",
             f"Histórico completo 2019 → {HOJE} · API Sienge · cada aba tem o detalhe parcela a parcela", len(COLS_R))
cabecalho(ws, 4, COLS_R, [44, 20, 8, 17, 17, 19], filtro=False)
r = 5
for k, apelido in enumerate(ORDEM):
    v, rf = D[apelido], refs[apelido]
    zebra = FILL_Z if k % 2 else None
    celula(ws, r, 1, v["nome"], fonte=F_B)
    celula(ws, r, 2, v["cnpj"], fonte=F_M)
    celula(ws, r, 3, rf["n"])
    celula(ws, r, 4, f"={rf['lanc']}", fonte=Font(name="Arial", size=9, color="008000"), fmt=MOEDA)
    celula(ws, r, 5, f"={rf.get('pago_sem') or rf['pago']}", fonte=Font(name="Arial", size=9, bold=True, color="008000"), fmt=MOEDA)
    if rf.get("aberto"):
        celula(ws, r, 6, f"={rf['aberto']}", fonte=Font(name="Arial", size=9, color="008000"), fmt=MOEDA)
    else:
        celula(ws, r, 6, "—", fonte=F_M, alinh=DIR)
    if zebra:
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = zebra
    ws.row_dimensions[r].height = 16
    r += 1
celula(ws, r + 1, 1, "MVCG é a representante de pagamento da Consplan: a linha soma as TEDs recebidas na conta dela (a mesma despesa está na linha da Consplan).", fonte=F_SUB)
celula(ws, r + 2, 1, "Verde = valor puxado por fórmula da aba da empresa · dica vermelha na célula = texto completo no comentário.", fonte=F_SUB)

destino = "output/Pagamentos_Fornecedores_Garden.xlsx"
try:
    wb.save(destino)
except PermissionError:
    destino = "output/Pagamentos_Fornecedores_Garden_novo.xlsx"
    try:
        wb.save(destino)
    except PermissionError:
        destino = "output/Pagamentos_Fornecedores_Garden_v2.xlsx"
        wb.save(destino)
    print("AVISO: arquivo aberto no Excel — salvei com outro nome")
print("gerado", destino)
