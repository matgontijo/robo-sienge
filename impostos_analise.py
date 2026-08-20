"""Cruza os titulos de imposto de 2026 com o padrao historico (2025 + 2026 dos demais usuarios)
e gera a lista de correcao. Fonte do tributo = CODIGO DE RECEITA lido na guia anexada.
Saidas: output/_correcao_impostos.json e output/CORRECAO_IMPOSTOS_2026.xlsx"""
import json, os, re
from datetime import datetime
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

H = json.load(open('output/_hist_impostos.json', encoding='utf-8'))
G = json.load(open('output/_codigos_guias.json', encoding='utf-8'))

COD = {'4095': 'RET', '1162': 'INSS', '5952': 'CSRF', '1708': 'IRRF', '8045': 'IRRF', '1732': 'ISS'}

# nomes dos planos financeiros (GET /payment-categories, salvo por impostos_historico.py)
PLANOS = {}
if os.path.exists('output/_planos_financeiros.json'):
    PLANOS = {p['id']: p['name'] for p in json.load(open('output/_planos_financeiros.json', encoding='utf-8'))}


def pmask(c):
    c = str(c)
    return f"{c[0]}.{c[1:3]}.{c[3:5]}.{c[5:]}" if len(c) == 7 else c


def pnome(c):
    """'2040303' -> '2.04.03.03 PIS/COFINS/CSLL' (ou so a mascara se o nome nao estiver no cache)."""
    n = PLANOS.get(str(c))
    return f"{pmask(c)} {n}" if n else pmask(c)


def tributo(t):
    cods = set()
    for gq in G.get(str(t['id']), {}).get('guias', []):
        for c, _ in gq['codigos']:
            cods.add(c)
    trib = {COD[c] for c in cods if c in COD}
    if len(trib) == 1:
        return trib.pop(), 'código de receita ' + '/'.join(sorted(c for c in cods if c in COD))
    txt = f"{t.get('documentNumber') or ''} {t.get('notes') or ''}".upper()
    tipo = (t.get('documentIdentificationId') or '').strip().upper()
    if 'IPTU' in txt or tipo == 'IPTU':
        return 'IPTU', 'texto'
    if re.search(r'\bINSS\b', txt):
        return 'INSS', 'texto'
    if re.search(r'\b(CSRF|CRF|CRSF)\b', txt):
        return 'CSRF', 'texto'
    if re.search(r'\b(IRRF|IFFR|IRF)\b', txt):
        return 'IRRF', 'texto'
    if re.search(r'\bISS\b', txt):
        return 'ISS', 'texto'
    if re.search(r'\bRET\b', txt):
        return 'RET', 'texto'
    if re.search(r'\b(TFE|TEO|TCLP|TAXA|TDE|TAXEXP|CND)\b', txt) or tipo in ('DAR', 'DARE', 'BOL', 'GUIA'):
        return 'TAXAS', 'texto'
    if tipo == 'FAT':
        return None, ''
    return 'OUTRO', 'texto'


def guia_valor(t):
    return [v for gq in G.get(str(t['id']), {}).get('guias', []) for v in gq['valores']]


def chave_aprop(t):
    ap = t['_apropriacoes']
    bd = t['_budget']
    obra = ' + '.join(sorted({f"{a['buildingName']} / {a['buildingUnitName']}" for a in ap})) or '—'
    plan = ' + '.join(f"{a['costEstimationSheetName']} ({a['percentage']:.0f}%)" if a['percentage'] != 100
                      else a['costEstimationSheetName'] for a in ap) or '—'
    cc = ' + '.join(sorted({str(b['costCenterId']) for b in bd})) or '—'
    plano = ' + '.join(sorted({b['paymentCategoriesId'] for b in bd})) or '—'
    return obra, plan, cc, plano


def plano_nome(plano):
    return ' + '.join(pnome(c) for c in plano.split(' + ')) if plano != '—' else '—'


CCN = {1: 'GARDEN - 1ª ETAPA', 2: 'GARDEN - 2ª ETAPA', 3: 'GARDEN - INCORPORAÇÃO', 100: 'ADMINISTRATIVO'}

# ---------- padrao historico por tributo ----------
hist = defaultdict(Counter)        # tributo -> (ano, obra, planilha, cc, plano) -> n
hist_val = defaultdict(float)
for t in H:
    trib, _ = tributo(t)
    if not trib or trib == 'OUTRO':
        continue
    ano = t['issueDate'][:4]
    if ano == '2026' and t['registeredBy'].startswith('MATHEUS'):
        continue   # o padrao e o dos outros
    obra, plan, cc, plano = chave_aprop(t)
    hist[trib][(ano, obra, plan, cc, plano)] += 1
    hist_val[(trib, ano, obra, plan, cc, plano)] += t['totalInvoiceAmount']

hist_sheets = defaultdict(set)   # tributo -> planilhas usadas no historico (2025/2026 outros)
hist_planos = defaultdict(set)
for t in H:
    trib, _ = tributo(t)
    if not trib or trib == 'OUTRO' or t['issueDate'] < '2025':
        continue
    if t['issueDate'][:4] == '2026' and t['registeredBy'].startswith('MATHEUS'):
        continue
    for a in t['_apropriacoes']:
        hist_sheets[trib].add(a['costEstimationSheetId'])
    for b in t['_budget']:
        hist_planos[trib].add(b['paymentCategoriesId'])

# padrao recomendado (2026 dos demais usuarios; 2025 como alternativa)
PADRAO = {
    'RET': dict(buildingId=3, buildingUnitId=600, obra='GARDEN - INCORPORAÇÃO / IMPOSTOS (RET)',
                sheetId='00.000.000.001', planilha='Imposto (RET)', cc=3, plano='2040503'),
    'INSS': dict(buildingId=1, buildingUnitId=2, obra='GARDEN - 1ª ETAPA / OBRA GARDEN',
                 sheetId='00.002.001.007', planilha='Equipe - Apoio (Até 12/2026)', cc=1, plano='1090107',
                 alt='2025: Remuneração PMG 1ª Etapa (Consplan / Cilicon) — 00.003.032.004'),
    'CSRF': dict(buildingId=1, buildingUnitId=2, obra='GARDEN - 1ª ETAPA / OBRA GARDEN',
                 sheetId='00.003.032.003', planilha='Taxa de Administração de Obra (Até 12/2026)', cc=1, plano='2040303',
                 alt='2025: Remuneração PMG 1ª Etapa — 00.003.032.004'),
    'IRRF': dict(buildingId=1, buildingUnitId=2, obra='GARDEN - 1ª ETAPA / OBRA GARDEN',
                 sheetId='00.003.032.004', planilha='Remuneração PMG 1ª Etapa (Consplan / Cilicon)', cc=1, plano='2040302',
                 alt='2.04.03.02 fica no grupo IMPOSTOS DE TERCEIROS, junto do ISS e PIS/COFINS/CSLL (Ludmylla/Thalita 2026); '
                     '2.04.02.06 tem o mesmo nome mas é do grupo de folha — foi o usado em 2025 e nos seus IRRF 06/07 (10446/10970)'),
    'ISS': dict(buildingId=1, buildingUnitId=2, obra='GARDEN - 1ª ETAPA / OBRA GARDEN',
                sheetId='00.003.032.003', planilha='Taxa de Administração de Obra (Até 12/2026)', cc=1, plano='2040301',
                alt='2025: Remuneração PMG 1ª Etapa — 00.003.032.004'),
    'IPTU': dict(buildingId=3, buildingUnitId=400, obra='GARDEN - INCORPORAÇÃO / TERRENO',
                 sheetId='01.000.000.003', planilha='IPTU/TLP', cc=3, plano='2040401'),
    'TAXAS': dict(buildingId=3, buildingUnitId=500, obra='GARDEN - INCORPORAÇÃO / INCORPORAÇÃO',
                  sheetId='04.000.000.004', planilha='Taxas', cc=3, plano='2010198',
                  alt='TEO (taxa de obra): GARDEN - 1ª ETAPA / Taxas de Legalização Emolumentos 00.001.001.013, CC 1'),
}
ORGAO_OK = ('RECEITA FEDERAL', 'DISTRITO FEDERAL', 'SECRETARIA DE ESTADO', 'FAZENDA', 'ECONOMIA')

# ---------- avaliacao dos titulos de 2026 ----------
linhas = []
for t in sorted(H, key=lambda x: (x['issueDate'], x['id'])):
    if t['issueDate'] < '2025':
        continue
    trib, fonte = tributo(t)
    if not trib or trib == 'OUTRO':
        continue
    obra, plan, cc, plano = chave_aprop(t)
    p = PADRAO[trib]
    problemas, acoes, observacoes = [], [], []
    obra_ids = {(a['buildingId'], a['buildingUnitId']) for a in t['_apropriacoes']}
    cc_ids = {b['costCenterId'] for b in t['_budget']}
    planos = {b['paymentCategoriesId'] for b in t['_budget']}
    sheets = {a['costEstimationSheetId'] for a in t['_apropriacoes']}
    erro_obra = bool(obra_ids) and obra_ids != {(p['buildingId'], p['buildingUnitId'])}
    erro_cc = bool(cc_ids) and cc_ids != {p['cc']}
    if trib == 'TAXAS':   # taxas tem dois destinos legitimos (incorporacao ou obra) - so aviso
        erro_obra = erro_cc = False
    if 'RPA' in (t.get('documentNumber') or '').upper():   # INSS de autonomo (RPA) segue o servico, nao a obra
        erro_obra = erro_cc = False
        problemas.append("INSS de RPA (autônomo): apropriado junto ao serviço (Marketing) — sem padrão comparável, provavelmente correto")
    if erro_obra:
        problemas.append(f"obra/unidade: está em {obra}; padrão {trib} = {p['obra']}")
        acoes.append(f"Apropriação de obra → {p['obra']} · {p['planilha']} ({p['sheetId']}) 100%")
    if erro_cc:
        nomes = ', '.join(CCN.get(int(c), c) for c in cc.split(' + ') if c.isdigit())
        problemas.append(f"centro de custo: está em {cc} ({nomes}); padrão {trib} = CC {p['cc']} ({CCN[p['cc']]})")
        acoes.append(f"Apropriação financeira → CC {p['cc']} ({CCN[p['cc']]}) · plano {pnome(p['plano'])} 100%")
    elif planos and p['plano'] not in planos and trib in ('RET', 'INSS', 'CSRF', 'ISS', 'IPTU', 'IRRF'):
        if planos <= hist_planos[trib]:
            observacoes.append(f"plano financeiro {plano_nome(plano)} (variação já usada no histórico; padrão 2026 = {pnome(p['plano'])})")
        else:
            problemas.append(f"plano financeiro: {plano_nome(plano)} nunca usado para {trib}; padrão = {pnome(p['plano'])}")
            acoes.append(f"Plano financeiro → {pnome(p['plano'])}")
    if not erro_obra and sheets and p['sheetId'] not in sheets and trib in ('INSS', 'CSRF', 'ISS', 'IRRF'):
        if sheets <= hist_sheets[trib]:
            observacoes.append(f"planilha {plan} (variação já usada no histórico; padrão 2026 = {p['planilha']})")
        else:
            problemas.append(f"planilha de custo: {plan} nunca usada para {trib}; padrão 2026 = {p['planilha']}")
    # credor
    if not any(o in t['_credor'].upper() for o in ORGAO_OK):
        problemas.append(f"credor: '{t['_credor']}' não é órgão arrecadador (padrão: SECRETARIA DA RECEITA FEDERAL / DISTRITO FEDERAL)")
        acoes.append("Credor → SECRETARIA DA RECEITA FEDERAL DO BRASIL (manual no Sienge; a API não altera credor)")
    # valor x guia
    gv = guia_valor(t)
    if gv and abs(sum(gv) - t['totalInvoiceAmount']) > 0.05 and abs(max(gv) - t['totalInvoiceAmount']) > 0.05:
        _br = lambda v: f"{v:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
        problemas.append(f"valor: título R$ {_br(t['totalInvoiceAmount'])} ≠ guia anexada R$ {_br(max(gv))} "
                         f"(dif. R$ {_br(t['totalInvoiceAmount'] - max(gv))})")
        acoes.append("Conferir valor pago x guia anexada (falta anexar guia complementar ou valor errado)")
    # nome do documento
    dn = (t.get('documentNumber') or '').upper()
    if trib == 'INSS' and 'RET' in dn:
        problemas.append("nome 'RET' no documento confunde com o RET (4095); a guia é INSS retido (1162)")
    if trib == 'CSRF' and 'RET' in dn:
        problemas.append("nome 'RET' no documento confunde com o RET (4095); a guia é CSRF (5952)")
    if trib == 'ISS' and not re.search(r'ISS', dn):
        problemas.append("nome do documento não diz ISS")
    if trib == 'INSS' and not re.search(r'\d{2}[-/]\d{4}', dn):
        problemas.append("nome do documento sem competência (mm/aaaa)")
    if re.search(r'[/-]2006\b', dn):
        problemas.append("nome do documento com ano 2006 (digitação)")

    status = 'CORRIGIR' if (erro_obra or erro_cc) else ('ATENÇÃO' if problemas else 'OK')
    linhas.append(dict(id=t['id'], data=t['issueDate'], tipo=(t['documentIdentificationId'] or '').strip(),
                       mes=t['issueDate'][:7], notas=(t.get('notes') or '')[:160],
                       doc=t.get('documentNumber') or '', valor=t['totalInvoiceAmount'], credor=t['_credor'],
                       quem=t['registeredBy'], tributo=trib, fonte=fonte, obra=obra, planilha=plan, cc=cc, plano=plano,
                       plano_nome=plano_nome(plano),
                       status=status, problemas=problemas, acoes=acoes, observacoes=observacoes, guia=gv, padrao=p,
                       erro_obra=bool(erro_obra), erro_cc=bool(erro_cc)))

for _k, _v in PADRAO.items():
    _v['plano_nome'] = pnome(_v['plano'])
json.dump({'gerado_em': datetime.now().isoformat(timespec='minutes'), 'n_titulos_lidos': len(H),
           'planos': {c: pnome(c) for c in PLANOS}, 'linhas': linhas, 'padrao': PADRAO,
           'historico': {trib: [dict(ano=k[0], obra=k[1], planilha=k[2], cc=k[3], plano=k[4], plano_nome=plano_nome(k[4]), n=n,
                                     total=round(hist_val[(trib,) + k], 2))
                                for k, n in c.most_common()] for trib, c in hist.items()}},
          open('output/_correcao_impostos.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

# ---------- resumo no terminal ----------
linhas_ano = [l for l in linhas if l['data'] >= '2026']
corr = [l for l in linhas_ano if l['status'] == 'CORRIGIR']
print(f"2026: {len(linhas_ano)} títulos de imposto | CORRIGIR {len(corr)} = R$ {sum(l['valor'] for l in corr):,.2f} | "
      f"ATENÇÃO {sum(1 for l in linhas_ano if l['status'] == 'ATENÇÃO')} | OK {sum(1 for l in linhas_ano if l['status'] == 'OK')}")
for l in linhas_ano:
    if l['status'] != 'OK':
        print(f"\n[{l['status']}] {l['id']} {l['tipo']} {l['doc']} R$ {l['valor']:,.2f} {l['tributo']} ({l['fonte']}) — {l['quem']}")
        for pbl in l['problemas']:
            print('   -', pbl)
        for a in l['acoes']:
            print('   →', a)

# ---------- XLSX ----------
wb = Workbook()
th = Font(bold=True, color='FFFFFF')
thf = PatternFill('solid', fgColor='16202A')
fill_err = PatternFill('solid', fgColor='FAEADC')
fill_at = PatternFill('solid', fgColor='FFF6D6')
fill_ok = PatternFill('solid', fgColor='E6F2EA')
thin = Side(style='thin', color='D8E0E4')
bd = Border(left=thin, right=thin, top=thin, bottom=thin)


def cab(ws, cols, widths):
    ws.append(cols)
    for i, w in enumerate(widths, 1):
        c = ws.cell(row=1, column=i)
        c.font = th
        c.fill = thf
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 30


ws = wb.active
ws.title = 'CORRIGIR'
cab(ws, ['Título', 'Emissão', 'Tipo', 'Documento', 'Valor', 'Tributo (pela guia)', 'Quem lançou',
         'ESTÁ EM — obra / unidade', 'ESTÁ EM — planilha', 'ESTÁ EM — CC', 'ESTÁ EM — plano fin.',
         'DEVE IR PARA — obra / unidade', 'DEVE IR PARA — planilha (código)', 'DEVE IR PARA — CC', 'DEVE IR PARA — plano fin.',
         'O que está errado', 'Ação'],
    [8, 11, 6, 24, 14, 12, 22, 34, 40, 8, 12, 34, 46, 8, 12, 70, 70])
for l in linhas_ano:
    if l['status'] == 'OK':
        continue
    p = l['padrao']
    ws.append([l['id'], l['data'], l['tipo'], l['doc'], l['valor'], f"{l['tributo']} — {l['fonte']}", l['quem'],
               l['obra'], l['planilha'], l['cc'], l['plano_nome'],
               p['obra'] if l['erro_obra'] else '(mantém)',
               f"{p['planilha']} ({p['sheetId']})" if l['erro_obra'] else '(mantém)',
               p['cc'] if l['erro_cc'] else '(mantém)', pnome(p['plano']) if l['erro_cc'] else '(mantém)',
               '\n'.join(l['problemas']), '\n'.join(l['acoes'])])
    r = ws.max_row
    for c in ws[r]:
        c.border = bd
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.fill = fill_err if l['status'] == 'CORRIGIR' else fill_at
    ws.cell(row=r, column=5).number_format = '#,##0.00'
tot = sum(l['valor'] for l in corr)
ws.append([])
ws.append(['', '', '', 'TOTAL A CORRIGIR (apropriação)', tot])
ws.cell(row=ws.max_row, column=5).number_format = '#,##0.00'
ws.cell(row=ws.max_row, column=4).font = Font(bold=True)
ws.cell(row=ws.max_row, column=5).font = Font(bold=True)

ws2 = wb.create_sheet('PADRÃO HISTÓRICO')
cab(ws2, ['Tributo', 'Ano', 'Obra / unidade', 'Planilha de custo', 'CC', 'Plano fin.', 'Qtde títulos', 'Total R$'],
    [10, 6, 34, 60, 6, 10, 12, 16])
ordem = ['RET', 'INSS', 'CSRF', 'IRRF', 'ISS', 'IPTU', 'TAXAS']
for trib in ordem:
    for k, n in sorted(hist[trib].items(), key=lambda kv: (-int(kv[0][0]), -kv[1])):
        if int(k[0]) < 2025:   # 2023-24 era outra rotina (DARFs pequenos de projetos); a base e 2025+
            continue
        ws2.append([trib, k[0], k[1], k[2], k[3], plano_nome(k[4]), n, round(hist_val[(trib,) + k], 2)])
        for c in ws2[ws2.max_row]:
            c.border = bd
            c.alignment = Alignment(vertical='top', wrap_text=True)
        ws2.cell(row=ws2.max_row, column=8).number_format = '#,##0.00'
ws2.append([])
ws2.append(['PADRÃO RECOMENDADO (2026, mesmo time):'])
ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True)
for trib in ordem:
    p = PADRAO[trib]
    ws2.append([trib, '', p['obra'], f"{p['planilha']} ({p['sheetId']})", p['cc'], pnome(p['plano']), '', p.get('alt', '')])

ws3 = wb.create_sheet('TODOS 2026')
cab(ws3, ['Título', 'Emissão', 'Tipo', 'Documento', 'Valor', 'Valor da guia', 'Tributo', 'Fonte', 'Credor', 'Quem lançou',
          'Obra / unidade', 'Planilha', 'CC', 'Plano fin.', 'Status', 'Observações'],
    [8, 11, 6, 24, 14, 14, 10, 22, 36, 22, 34, 44, 6, 12, 10, 70])
for l in linhas_ano:
    ws3.append([l['id'], l['data'], l['tipo'], l['doc'], l['valor'], max(l['guia']) if l['guia'] else None,
                l['tributo'], l['fonte'], l['credor'], l['quem'], l['obra'], l['planilha'], l['cc'], l['plano_nome'],
                l['status'], '\n'.join(l['problemas'] + l['observacoes'])])
    r = ws3.max_row
    for c in ws3[r]:
        c.border = bd
        c.alignment = Alignment(vertical='top', wrap_text=True)
        c.fill = {'CORRIGIR': fill_err, 'ATENÇÃO': fill_at, 'OK': fill_ok}[l['status']]
    ws3.cell(row=r, column=5).number_format = '#,##0.00'
    ws3.cell(row=r, column=6).number_format = '#,##0.00'
try:
    wb.save('output/CORRECAO_IMPOSTOS_2026.xlsx')
    print('\nsalvo output/CORRECAO_IMPOSTOS_2026.xlsx')
except PermissionError:   # planilha aberta no Excel: nao derruba o resto da analise
    wb.save('output/CORRECAO_IMPOSTOS_2026_novo.xlsx')
    print('\nAVISO: CORRECAO_IMPOSTOS_2026.xlsx esta aberta no Excel; salvei como CORRECAO_IMPOSTOS_2026_novo.xlsx')


# ---------- plano do painel de correcao (/apropriacao) ----------
OPCOES = {
    "INSS": [dict(sheet="00.002.001.007", planilha="Equipe - Apoio (Até 12/2026)", plano="1090107", rotulo="padrão 2026 (Thalita/Ludmylla e seu INSS 06)"),
             dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="1090107", rotulo="padrão 2025 / jan-mar 2026")],
    "CSRF": [dict(sheet="00.003.032.003", planilha="Taxa de Administração de Obra (Até 12/2026)", plano="2040303", rotulo="padrão 2026 (e seu CSRF 07)"),
             dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="2040303", rotulo="padrão 2025 / jan-mar 2026")],
    "ISS": [dict(sheet="00.003.032.003", planilha="Taxa de Administração de Obra (Até 12/2026)", plano="2040301", rotulo="padrão 2026"),
            dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="2040301", rotulo="padrão 2025"),
            dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="1090104", rotulo="Ludmylla jan-fev 2026")],
    "IRRF": [dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="2040302", rotulo="grupo Impostos de Terceiros, como ISS/CSRF (Ludmylla/Thalita 2026)"),
             dict(sheet="00.003.032.004", planilha="Remuneração PMG 1ª Etapa (Consplan / Cilicon)", plano="2040206", rotulo="mesmo nome, grupo de folha — igual aos seus IRRF 06 e 07")],
}
for _ops in OPCOES.values():
    for _o in _ops:
        _o['plano_nome'] = pnome(_o['plano'])
PLANO_PATH = 'output/_plano_apropriacao.json'
antigo = json.load(open(PLANO_PATH, encoding='utf-8')) if os.path.exists(PLANO_PATH) else {}
escolhas = {i['id']: i.get('escolha', 0) for i in antigo.get('itens', [])}
extra = antigo.get('manual_extra', ["RET 02/2026 e RET 05/2026 não localizados no contas a pagar"])
itens_plano = []
for l in linhas:
    if l['status'] != 'CORRIGIR' or l['tributo'] not in OPCOES:
        continue
    cod = sorted({c for gq in G.get(str(l['id']), {}).get('guias', []) for c, _ in gq['codigos']})
    itens_plano.append(dict(id=l['id'], tipo=l['tipo'], doc=l['doc'], data=l['data'], valor=l['valor'], tributo=l['tributo'],
                            codigo='/'.join(cod), credor=l['credor'], quem=l['quem'],
                            de=dict(obra=l['obra'], planilha=l['planilha'], cc=l['cc'], plano=l['plano'], plano_nome=l['plano_nome']),
                            escolha=escolhas.get(l['id'], 0),
                            avisos=[p for p in l['problemas'] if not p.startswith(('obra/unidade', 'centro de custo'))]))
manual = []
for l in linhas:
    for p in l['problemas']:
        if p.startswith('credor:'):
            manual.append(dict(id=l['id'], oque=f"Credor: {l['credor'][:28]}… → SECRETARIA DA RECEITA FEDERAL DO BRASIL"))
        if p.startswith('valor:'):
            manual.append(dict(id=l['id'], oque="V" + p[1:] + " — conferir/anexar guia complementar"))
manual += [dict(id=None, oque=m) for m in extra]
json.dump(dict(obra=dict(buildingId=1, buildingUnitId=2, nome="GARDEN - 1ª ETAPA / OBRA GARDEN"), cc=dict(id=1, nome="GARDEN - 1ª ETAPA"),
               opcoes=OPCOES, itens=itens_plano, manual=manual, manual_extra=extra, planos={c: pnome(c) for c in PLANOS},
               gerado_em=datetime.now().isoformat(timespec='minutes'),
               base="2025 + 2026 (Ludmylla/Thalita) — tributo pelo código de receita da guia anexada"),
          open(PLANO_PATH, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
print(f"plano do painel: {len(itens_plano)} títulos a corrigir")
