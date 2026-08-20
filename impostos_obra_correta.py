"""Pelos DADOS, qual e a obra/planilha correta de cada guia de retencao (INSS/CSRF/IRRF/ISS)?
Regra: a retencao segue a nota que a gerou. Para cada guia, soma os impostos retidos das NFs da
competencia (GET /bills/{id}/taxes, ja no livro output/_retencoes_nf_2026.json) e distribui o valor da
guia pela apropriacao de obra dessas NFs (buildings-cost). Confere com a guia (total e, no INSS, por CNPJ).
Saidas: output/_obra_correta.json e output/OBRA_CORRETA_IMPOSTOS_2026.xlsx"""
import json
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

LED = json.load(open('output/_retencoes_nf_2026.json', encoding='utf-8'))
H = {t['id']: t for t in json.load(open('output/_hist_impostos.json', encoding='utf-8'))}
COMP = json.load(open('output/_darf_composicao.json', encoding='utf-8'))
C = json.load(open('output/_correcao_impostos.json', encoding='utf-8'))
LIN = {l['id']: l for l in C['linhas']}
CRED = json.load(open('output/_credores_cache.json', encoding='utf-8'))
PLANOS = {p['id']: p['name'] for p in json.load(open('output/_planos_financeiros.json', encoding='utf-8'))}

# que taxId do Sienge compoe cada guia
def trib_do_tax(taxid):
    t = (taxid or '').upper()
    if 'INSS' in t: return 'INSS'
    if 'ISS' in t: return 'ISS'
    if 'CSRF' in t or 'PIS' in t or 'COFINS' in t or 'CSLL' in t: return 'CSRF'
    if 'IR' in t: return 'IRRF'
    return None

def mes(iso): return iso[:7]
def mes_ant(m):
    a, mm = int(m[:4]), int(m[5:7]); mm -= 1
    if mm == 0: a, mm = a - 1, 12
    return f"{a}-{mm:02d}"
def fmt_aprop(aprop):
    """'Planilha A (30%) + Planilha B' a partir da lista de apropriacoes."""
    partes = []
    for ap in aprop or []:
        nome = (ap.get('costEstimationSheetName') or '').strip()
        pct = ap.get('percentage') or 0
        partes.append(nome if pct == 100 else f"{nome} ({pct:.0f}%)")
    return ' + '.join(partes) or '—'


def so_dig(s): return re.sub(r'\D', '', s or '')

# indice: (tributo, mes da NF) -> linhas [nf, amount]
idx = defaultdict(list)
for e in LED.values():
    for tx in e.get('taxes', []):
        tr = trib_do_tax(tx['tax'])
        if tr and tx['amount']:
            idx[(tr, mes(e['issueDate']))].append((e, tx['amount']))

# CNPJ por credor (para conferir o INSS por prestador)
cnpj_cred = {}
import requests, config
from requests.auth import HTTPBasicAuth
S = requests.Session(); S.auth = HTTPBasicAuth(config.SIENGE_USERNAME, config.SIENGE_PASSWORD)
def cnpj_do_credor(cid):
    cid = str(cid)
    if cid not in cnpj_cred:
        try:
            r = S.get(f"{config.SIENGE_BASE_URL}/creditors/{cid}", timeout=60).json()
            cnpj_cred[cid] = so_dig(r.get('cnpj') or r.get('cpf') or '')
        except Exception:
            cnpj_cred[cid] = ''
    return cnpj_cred[cid]

resultados = []
for tid, t in sorted(H.items(), key=lambda kv: kv[1]['issueDate']):
    if t['issueDate'] < '2026' or tid not in LIN: continue
    l = LIN[tid]; trib = l['tributo']
    if trib not in ('INSS', 'CSRF', 'IRRF', 'ISS'): continue
    comp = COMP.get(str(tid), [])
    pas = sorted({c['pa'] for c in comp if c.get('pa')})
    if pas:
        m_comp = f"{pas[0][3:]}-{pas[0][:2]}"
    else:
        mm = re.search(r'(\d{2})[-/](\d{4})', t['documentNumber'] or '')
        m_comp = f"{mm.group(2)}-{mm.group(1)}" if mm else mes(t['issueDate'])
    # testa NFs do mes da competencia e do mes anterior: fica com o que mais bate com a guia
    melhor = None
    for m_nf, rot in ((m_comp, 'NFs emitidas no mês da competência'), (mes_ant(m_comp), 'NFs emitidas no mês anterior')):
        linhas = idx.get((trib, m_nf), [])
        soma = sum(a for _, a in linhas)
        dif = abs(soma - t['totalInvoiceAmount'])
        if melhor is None or dif < melhor['dif']:
            melhor = dict(m_nf=m_nf, rot=rot, linhas=linhas, soma=soma, dif=dif)
    linhas, soma = melhor['linhas'], melhor['soma']
    # rateio por planilha (peso = imposto retido x % da apropriacao da NF)
    rateio = defaultdict(float); por_credor = defaultdict(float); nfs = []
    for e, a in linhas:
        aprop = e.get('aprop') or []
        if not aprop:
            rateio[('?', '?', '?', 'NF sem apropriação', None)] += a
        for ap in aprop:
            rateio[(ap['buildingId'], ap['buildingUnitName'], ap['costEstimationSheetId'], ap['costEstimationSheetName'].strip(), ap['buildingUnitId'])] += a * (ap['percentage'] or 0) / 100
        por_credor[e.get('credor') or str(e['creditorId'])] += a
        nfs.append(dict(id=e['id'], doc=f"{e['tipo']} {e['doc']}", credor=e.get('credor', ''), data=e['issueDate'], imposto=round(a, 2),
                        planilha=fmt_aprop(aprop)))
    if not linhas:
        rateio[('?', '?', '?', 'sem NF no livro das retenções para esse mês', None)] = t['totalInvoiceAmount']
    tot_r = sum(rateio.values()) or 1
    rateio_l = sorted([dict(obra=k[0], unidade=k[1], sheet=k[2], planilha=k[3], unidade_id=k[4], valor=round(v, 2), pct=round(100 * v / tot_r, 2))
                       for k, v in rateio.items()], key=lambda x: -x['valor'])
    # conferencia por CNPJ (INSS)
    conf_cnpj = None
    if trib == 'INSS' and comp:
        darf = defaultdict(float)
        for c in comp:
            if c.get('cnpj'): darf[so_dig(c['cnpj'])] += c['principal']
        led = defaultdict(float)
        for e, a in linhas:
            led[cnpj_do_credor(e['creditorId'])] += a
        bate = sum(1 for k, v in darf.items() if abs(led.get(k, 0) - v) < 0.05)
        so_darf = [(CRED.get(k, k), round(v, 2)) for k, v in darf.items() if k not in led]
        so_led = [(k, round(v, 2)) for k, v in led.items() if k not in darf]
        conf_cnpj = dict(prestadores_darf=len(darf), batem=bate,
                         so_no_darf=[(next((n for c_, n in CRED.items() if so_dig(cnpj_cred.get(c_, '')) == k), k), v) for k, v in so_darf],
                         so_nas_nfs=[(CRED.get(next((c_ for c_, cn in cnpj_cred.items() if cn == k), ''), k), v) for k, v in so_led])
    atual = fmt_aprop(t['_apropriacoes'])
    resultados.append(dict(id=tid, doc=f"{(t['documentIdentificationId'] or '').strip()} {t['documentNumber']}", valor=t['totalInvoiceAmount'],
                           tributo=trib, quem=t['registeredBy'].title(), competencia=m_comp, mes_nf=melhor['m_nf'], regra=melhor['rot'],
                           soma_nfs=round(soma, 2), dif=round(t['totalInvoiceAmount'] - soma, 2), n_nfs=len(linhas),
                           atual=atual, atual_obra=' + '.join(sorted({f"{a['buildingName']} / {a['buildingUnitName']}" for a in t['_apropriacoes']})),
                           rateio=rateio_l, por_credor=sorted([(k, round(v, 2)) for k, v in por_credor.items()], key=lambda x: -x[1]),
                           nfs=sorted(nfs, key=lambda x: -x['imposto']), conf_cnpj=conf_cnpj))

json.dump(resultados, open('output/_obra_correta.json', 'w', encoding='utf-8'), ensure_ascii=False, indent=1)

# ---------- terminal ----------
for r in resultados:
    flag = 'OK' if abs(r['dif']) < max(1.0, 0.005 * r['valor']) else f"DIF {r['dif']:,.2f}"
    print(f"\n{r['id']} {r['doc'][:24]:24} R$ {r['valor']:>11,.2f} [{r['tributo']}] comp {r['competencia']} | NFs de {r['mes_nf']}: {r['n_nfs']} notas somam {r['soma_nfs']:,.2f} -> {flag} | {r['quem']}")
    print(f"   hoje: {r['atual_obra'][:40]} · {r['atual'][:70]}")
    for x in r['rateio'][:6]:
        print(f"   {x['pct']:5.1f}%  {x['planilha'][:55]:55} ({x['unidade']})  R$ {x['valor']:,.2f}")
    if r['conf_cnpj']:
        cc = r['conf_cnpj']; print(f"   conferência por CNPJ (INSS): {cc['batem']}/{cc['prestadores_darf']} prestadores batem; só no DARF: {cc['so_no_darf'][:3]}; só nas NFs: {cc['so_nas_nfs'][:3]}")

# ---------- XLSX ----------
def acerto_atual(r):
    """% do valor da guia que a apropriacao ATUAL coloca na planilha certa (intersecao com o rateio das NFs)."""
    cur = defaultdict(float)
    for a in H[r['id']]['_apropriacoes']:
        cur[(a['buildingId'], a['costEstimationSheetId'])] += (a['percentage'] or 0)
    return round(sum(min(x['pct'], cur.get((x['obra'], x['sheet']), 0)) for x in r['rateio']), 1)


wb = Workbook(); ws0 = wb.active; ws0.title = 'RESUMO'
ws0.append(['Título', 'Documento', 'Tributo', 'Competência', 'Quem lançou', 'Valor da guia', 'NFs de origem', 'Soma das NFs', 'Diferença', 'Está hoje em', 'Acerto da apropriação atual (%)', 'Planilhas corretas (qtde)', 'Maior planilha pelas NFs', '% da maior'])
for i, w in enumerate([8, 24, 8, 11, 22, 14, 10, 14, 12, 44, 14, 12, 44, 10], 1):
    c = ws0.cell(row=1, column=i); c.font = Font(bold=True, color='FFFFFF'); c.fill = PatternFill('solid', fgColor='16202A'); ws0.column_dimensions[get_column_letter(i)].width = w
ws0.freeze_panes = 'A2'
for r in resultados:
    confiavel = abs(r['dif']) <= max(2.0, 0.01 * r['valor'])
    top = r['rateio'][0] if r['rateio'] else {}
    ws0.append([r['id'], r['doc'], r['tributo'], r['competencia'], r['quem'], r['valor'], r['n_nfs'], r['soma_nfs'], r['dif'], r['atual'],
                acerto_atual(r) if confiavel else None, len(r['rateio']) if confiavel else None, top.get('planilha'), top.get('pct')])
    for col in (6, 8, 9): ws0.cell(row=ws0.max_row, column=col).number_format = '#,##0.00'
ws = wb.create_sheet('RATEIO CORRETO')
th = Font(bold=True, color='FFFFFF'); thf = PatternFill('solid', fgColor='16202A'); thin = Side(style='thin', color='D8E0E4'); bd = Border(left=thin, right=thin, top=thin, bottom=thin)
cols = ['Título', 'Documento', 'Tributo', 'Valor da guia', 'Competência', 'NFs usadas', 'Soma das NFs', 'Diferença', 'Está hoje em', 'Obra / unidade (pelas NFs)', 'Planilha (pelas NFs)', '% ', 'R$']
ws.append(cols)
for i, w in enumerate([8, 24, 8, 14, 11, 10, 14, 12, 40, 30, 52, 8, 14], 1):
    c = ws.cell(row=1, column=i); c.font = th; c.fill = thf; ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = 'A2'
for r in resultados:
    first = True
    for x in r['rateio']:
        ws.append([r['id'], r['doc'], r['tributo'], r['valor'], r['competencia'], r['n_nfs'], r['soma_nfs'], r['dif'], r['atual'],
                   f"{x['obra']} / {x['unidade']}", f"{x['planilha']} ({x['sheet']})", x['pct'], x['valor']] if first else
                  ['', '', '', None, '', None, None, None, '', f"{x['obra']} / {x['unidade']}", f"{x['planilha']} ({x['sheet']})", x['pct'], x['valor']])
        for c in ws[ws.max_row]: c.border = bd; c.alignment = Alignment(vertical='top', wrap_text=True)
        for col in (4, 7, 8, 13): ws.cell(row=ws.max_row, column=col).number_format = '#,##0.00'
        ws.cell(row=ws.max_row, column=12).number_format = '0.0'
        first = False
ws2 = wb.create_sheet('NFs DE ORIGEM')
ws2.append(['Guia', 'Tributo', 'NF', 'Documento', 'Credor', 'Emissão', 'Imposto retido', 'Planilha da NF'])
for i, w in enumerate([8, 8, 8, 18, 44, 11, 14, 60], 1):
    c = ws2.cell(row=1, column=i); c.font = th; c.fill = thf; ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'
for r in resultados:
    for n in r['nfs']:
        ws2.append([r['id'], r['tributo'], n['id'], n['doc'], n['credor'], n['data'], n['imposto'], n['planilha']])
        ws2.cell(row=ws2.max_row, column=7).number_format = '#,##0.00'
wb.save('output/OBRA_CORRETA_IMPOSTOS_2026.xlsx')
print('\nsalvo output/OBRA_CORRETA_IMPOSTOS_2026.xlsx e output/_obra_correta.json')


# ---------- grava o rateio no plano do painel (/apropriacao): opcao "rateio pelas NFs de origem" ----------
PLANO_PATH = 'output/_plano_apropriacao.json'
import os as _os
plano = json.load(open(PLANO_PATH, encoding='utf-8')) if _os.path.exists(PLANO_PATH) else None
if plano is not None:
    PLANO_TRIB = {'INSS': '1090107', 'CSRF': '2040303', 'ISS': '2040301', 'IRRF': '2040302'}
    CC_DA_OBRA = {1: 1, 2: 2, 3: 3, 4: 4}
    por_id = {it['id']: it for it in plano['itens']}
    for r in resultados:
        linhas_ok = [x for x in r['rateio'] if x['obra'] != '?' and x['unidade_id'] is not None]
        if not linhas_ok or abs(r['dif']) > max(2.0, 0.01 * r['valor']):
            continue   # sem base confiavel: nao oferece rateio
        # percentuais com 2 casas somando exatamente 100
        tot = sum(x['valor'] for x in linhas_ok)
        pcts = [round(100 * x['valor'] / tot, 2) for x in linhas_ok]
        pcts[0] = round(pcts[0] + (100 - sum(pcts)), 2)
        obra_l = [dict(buildingId=x['obra'], buildingUnitId=x['unidade_id'], unidade=x['unidade'], sheet=x['sheet'], planilha=x['planilha'], pct=pc, valor=x['valor'])
                  for x, pc in zip(linhas_ok, pcts) if pc > 0]
        # CC segue a obra; plano financeiro segue o tributo
        cc_acc = defaultdict(float)
        for x in obra_l:
            cc_acc[CC_DA_OBRA.get(x['buildingId'], x['buildingId'])] += x['pct']
        cc_l = [dict(costCenterId=k, plano=PLANO_TRIB[r['tributo']], pct=round(v, 2)) for k, v in cc_acc.items()]
        cc_l[0]['pct'] = round(cc_l[0]['pct'] + (100 - sum(c['pct'] for c in cc_l)), 2)
        _br = lambda v: f"{v:,.2f}".replace(",", "@").replace(".", ",").replace("@", ".")
        resumo = f"{r['n_nfs']} NFs de {r['mes_nf'][5:]}/{r['mes_nf'][:4]} somam R$ {_br(r['soma_nfs'])} (guia R$ {_br(r['valor'])}); maior: {obra_l[0]['planilha'][:40]} {obra_l[0]['pct']}%"
        rateio_nf = dict(obra=obra_l, cc=cc_l, resumo=resumo, plano_nome=PLANOS.get(PLANO_TRIB[r['tributo']], ''))
        t = H[r['id']]
        if r['id'] in por_id:
            por_id[r['id']]['rateio_nf'] = rateio_nf
        else:
            l = LIN[r['id']]
            plano['itens'].append(dict(id=r['id'], tipo=(t['documentIdentificationId'] or '').strip(), doc=t['documentNumber'], data=t['issueDate'],
                                       valor=t['totalInvoiceAmount'], tributo=r['tributo'], codigo=l['fonte'].replace('código de receita ', ''),
                                       credor=t['_credor'], quem=t['registeredBy'],
                                       de=dict(obra=l['obra'], planilha=l['planilha'], cc=l['cc'], plano=l['plano'], plano_nome=l.get('plano_nome', l['plano'])),
                                       escolha=0, avisos=[], rateio_nf=rateio_nf))
    plano['itens'].sort(key=lambda it: (it['data'], it['id']))
    plano['base'] = "rateio pelas NFs de origem (GET /bills/{id}/taxes + buildings-cost) · variantes da equipe 2025-2026"
    json.dump(plano, open(PLANO_PATH, 'w', encoding='utf-8'), ensure_ascii=False, indent=1)
    print(f"plano do painel: {len(plano['itens'])} guias com opção de rateio pelas NFs")
