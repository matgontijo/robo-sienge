import os
from datetime import date
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from loguru import logger
from models import Titulo
from modules.reconciler import Divergencia

class ReportGenerator:
    def __init__(self):
        self.fill_header = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        self.fill_critica = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        self.fill_atencao = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.fill_ok = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.font_bold = Font(bold=True)
        self.font_link = Font(color="0563C1", underline="single")

    def gerar(
        self,
        divergencias: List[Tuple[Titulo, List[Divergencia]]],
        titulos_ok: List[Titulo],
        titulos_erro: List[Tuple[Titulo, str]], # Titulo e motivo do erro
        data_referencia: date,
        output_dir: str,
        pagamentos: List[dict] = None,
        dossie: List[dict] = None
    ) -> str:

        wb = Workbook()

        # Aba Divergncias
        ws_div = wb.active
        ws_div.title = "Divergências"
        self._preencher_aba_divergencias(ws_div, divergencias)

        # Aba Pagamentos (destino cadastrado na parcela x CNPJ do credor)
        if pagamentos:
            ws_pag = wb.create_sheet(title="Pagamentos")
            self._preencher_aba_pagamentos(ws_pag, pagamentos)

        # Aba Dossiê (checklist dos documentos essenciais copiados p/ pasta única)
        if dossie:
            ws_dos = wb.create_sheet(title="Dossiê")
            self._preencher_aba_dossie(ws_dos, dossie)

        # Aba Conferidos OK
        ws_ok = wb.create_sheet(title="Conferidos OK")
        self._preencher_aba_ok(ws_ok, titulos_ok)

        # Aba No Processados
        ws_err = wb.create_sheet(title="Não Processados")
        self._preencher_aba_erro(ws_err, titulos_erro)
        
        # Salvar
        os.makedirs(os.path.join(output_dir, "relatorios"), exist_ok=True)
        timestamp = data_referencia.strftime("%Y-%m-%d")
        
        import datetime
        agora = datetime.datetime.now().strftime("%H%M")
        
        filename = f"conciliacao_{timestamp}_{agora}.xlsx"
        filepath = os.path.join(output_dir, "relatorios", filename)
        
        wb.save(filepath)
        logger.success(f"Relatrio gerado: {filepath}")
        return filepath

    def _preencher_aba_divergencias(self, ws, divergencias_list):
        headers = [
            "Nº Título Sienge", "Fornecedor", "CNPJ Fornecedor", "Valor Sienge", 
            "Vencimento Sienge", "Forma Pagamento", "Tipo Divergência", 
            "Campo", "Valor Sienge (D)", "Valor NF-e", "Valor Boleto", 
            "Criticidade", "Link DANFE"
        ]
        ws.append(headers)
        
        # Formatar cabealho
        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_bold
            
        row_num = 2
        qtd_criticas = 0
        qtd_atencao = 0
        
        for titulo, divs in divergencias_list:
            for div in divs:
                ws.append([
                    titulo.numero,
                    titulo.fornecedor_nome,
                    titulo.fornecedor_cnpj,
                    titulo.valor_liquido,
                    str(titulo.data_vencimento),
                    titulo.forma_pagamento,
                    div.tipo,
                    div.campo,
                    div.valor_sienge,
                    div.valor_nfe,
                    div.valor_boleto,
                    div.criticidade,
                    div.danfe_path or "Nenhum"
                ])
                
                # Pintar linha baseada na criticidade
                fill = None
                if div.criticidade == "CRITICA":
                    fill = self.fill_critica
                    qtd_criticas += 1
                elif div.criticidade == "ATENCAO":
                    fill = self.fill_atencao
                    qtd_atencao += 1
                    
                if fill:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_num, column=col).fill = fill
                
                # Transformar Link DANFE em link se existir
                if div.danfe_path:
                    cell_link = ws.cell(row=row_num, column=13)
                    # Caminho absoluto para o link funcionar localmente
                    abs_path = os.path.abspath(div.danfe_path)
                    cell_link.hyperlink = f"file:///{abs_path.replace(chr(92), '/')}"
                    cell_link.font = self.font_link
                
                row_num += 1
                
        # Totais
        ws.append([])
        ws.append(["TOTAIS", f"Críticas: {qtd_criticas}", f"Atenção: {qtd_atencao}", f"Total Divergências: {qtd_criticas + qtd_atencao}"])
        for cell in ws[row_num+1]:
            cell.font = self.font_bold
            
        # Filtro
        ws.auto_filter.ref = f"A1:M{row_num - 1}"
        
        # Auto ajuste
        self._auto_fit_columns(ws)

    def _preencher_aba_pagamentos(self, ws, pagamentos):
        """Dados de pagamento da parcela (aba Inf. Pagamento do Sienge) e o
        confronto do destino (chave Pix / conta TED) com o CNPJ do credor."""
        headers = [
            "Nº Título", "Parcela", "Fornecedor", "CNPJ Credor (cadastro)",
            "Forma Pagamento", "Valor Parcela", "Vencimento",
            "Tipo Chave Pix", "Chave Pix", "Banco", "Agência", "Conta",
            "Titular/Beneficiário", "CNPJ Destino", "Linha Digitável",
            "Banco do Boleto", "Valor Boleto (linha)",
            "Confronto CNPJ"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_bold

        row_num = 2
        for p in pagamentos:
            ws.append([
                p.get("numero"), p.get("parcela"), p.get("fornecedor"),
                p.get("cnpj_credor"), p.get("forma"), p.get("valor"),
                str(p.get("vencimento") or ""), p.get("tipo_chave_pix"),
                p.get("chave_pix"), p.get("banco"), p.get("agencia"),
                p.get("conta"), p.get("titular"), p.get("cnpj_destino"),
                p.get("linha_digitavel"), p.get("banco_boleto"),
                p.get("valor_boleto"), p.get("confronto"),
            ])
            confronto = str(p.get("confronto") or "")
            fill = None
            if confronto == "DIVERGENTE":
                fill = self.fill_critica
            elif confronto == "OK":
                fill = self.fill_ok
            elif confronto.startswith("NAO"):
                fill = self.fill_atencao
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill
            row_num += 1

        ws.auto_filter.ref = f"A1:R{row_num - 1}"
        self._auto_fit_columns(ws)

    def _preencher_aba_dossie(self, ws, dossie):
        """Checklist do dossiê: cada título com o status da NF e do boleto
        (✓ copiado p/ a pasta única · FALTA · n/a quando não se aplica)."""
        headers = ["Nº Título", "Parcela", "Fornecedor", "Tipo Doc", "Forma Pagto",
                   "NF", "Boleto", "Arquivo NF", "Arquivo Boleto", "Anexos no título"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_bold

        row_num = 2
        faltas_nf = faltas_bol = 0
        for d in dossie:
            ws.append([
                d.get("numero"), d.get("parcela"), d.get("fornecedor"),
                d.get("tipo_doc"), d.get("forma"), d.get("nf"), d.get("boleto"),
                d.get("arq_nf") or "", d.get("arq_boleto") or "", d.get("total_anexos"),
            ])
            fill = None
            if d.get("nf") == "FALTA":
                fill = self.fill_critica
                faltas_nf += 1
            if d.get("boleto") == "FALTA":
                fill = fill or self.fill_atencao
                faltas_bol += 1
            if d.get("nf") == "✓" and d.get("boleto") in ("✓", "n/a") and fill is None:
                fill = self.fill_ok
            if fill:
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_num, column=col).fill = fill
            row_num += 1

        ws.append([])
        ws.append(["TOTAIS", f"Sem NF: {faltas_nf}", f"Sem boleto: {faltas_bol}",
                   f"Completos: {sum(1 for d in dossie if d.get('nf') == '✓' and d.get('boleto') in ('✓', 'n/a'))}"])
        for cell in ws[row_num + 1]:
            cell.font = self.font_bold
        ws.auto_filter.ref = f"A1:J{row_num - 1}"
        self._auto_fit_columns(ws)

    def _preencher_aba_ok(self, ws, titulos):
        headers = [
            "Nº Título", "Fornecedor", "CNPJ", "Valor", 
            "Vencimento", "Pagamento"
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_bold
            
        for t in titulos:
            ws.append([
                t.numero, t.fornecedor_nome, t.fornecedor_cnpj,
                t.valor_liquido, str(t.data_vencimento), t.forma_pagamento
            ])
            
        ws.auto_filter.ref = f"A1:F{len(titulos)+1}"
        self._auto_fit_columns(ws)

    def _preencher_aba_erro(self, ws, titulos_erro):
        headers = ["Nº Título", "Fornecedor", "Valor", "Erro/Motivo"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = self.fill_header
            cell.font = self.font_bold
            
        for t, erro in titulos_erro:
            ws.append([
                t.numero, t.fornecedor_nome, t.valor_liquido, erro
            ])
            
        ws.auto_filter.ref = f"A1:D{len(titulos_erro)+1}"
        self._auto_fit_columns(ws)

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = min(adjusted_width, 50) # limite 50
