import os
import re
import unicodedata
from datetime import date, datetime
from typing import List, Optional, Tuple

from loguru import logger

from models import Titulo

# Valores válidos para a coluna "Orig."
ORIGENS_VALIDAS = {"AC", "CP", "ME", "CO"}

# Mapeamento de rótulos normalizados do cabeçalho -> chave interna
_HEADER_KEYS = {
    "data": "data",
    "documento": "documento",
    "tit/parc": "tit_parc",
    "tit / parc": "tit_parc",
    "tit parc": "tit_parc",
    "orig": "orig",
    "cliente/fornecedor": "fornecedor",
    "cliente / fornecedor": "fornecedor",
    "entradas": "entradas",
    "saidas": "saidas",
    "saldo": "saldo",
}


def _normalizar(texto: object) -> str:
    """Minúsculas, sem acento, sem pontuação de borda — para casar rótulos."""
    if texto is None:
        return ""
    s = str(texto).strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return s.strip(" .:")


class ReportParser:
    """
    Lê um relatório de Fluxo de Caixa Analítico exportado do Sienge (.xlsx ou .pdf)
    e devolve a lista de títulos *conferíveis* (apenas saídas/pagamentos).

    Regras aplicadas:
    - Ignora linhas de "Total do dia/período", cabeçalhos repetidos e linhas em branco.
    - Considera apenas linhas com Saídas > 0 (pagamentos). Linhas com Entradas > 0
      (créditos de cliente) são descartadas.
    - Documento `TIPO.NUMERO` é separado em tipo_documento/numero_documento.
    - Tit/Parc `NUMERO/PARCELA` é separado em numero (título) e parcela.
    """

    # ------------------------------------------------------------------ API

    def parse(self, path: str) -> List[Titulo]:
        ext = os.path.splitext(path)[1].lower()
        logger.info(f"Lendo relatório de Fluxo de Caixa: {path} (ext={ext})")
        if ext in (".xlsx", ".xlsm"):
            return self._parse_xlsx(path)
        if ext == ".pdf":
            return self._parse_pdf(path)
        raise ValueError(f"Extensão de relatório não suportada: {ext} (use .xlsx ou .pdf)")

    # ------------------------------------------------------------- helpers

    @staticmethod
    def _parse_valor_br(valor: object) -> float:
        """Converte valores BR ('1.234,56', '(1.234,56)', 'R$ 10,00') para float."""
        if valor is None:
            return 0.0
        if isinstance(valor, (int, float)):
            return float(valor)
        s = str(valor).strip()
        if not s or s in ("-", "--"):
            return 0.0
        negativo = s.startswith("(") and s.endswith(")")
        s = s.replace("R$", "").replace("(", "").replace(")", "").strip()
        s = s.replace(".", "").replace(",", ".")
        s = re.sub(r"[^0-9.\-]", "", s)
        if not s or s in (".", "-"):
            return 0.0
        try:
            v = float(s)
        except ValueError:
            return 0.0
        return -v if negativo else v

    @staticmethod
    def _limpar_texto(valor: object) -> str:
        """Colapsa quebras de linha e espaços repetidos (células quebradas no PDF)."""
        if valor is None:
            return ""
        return " ".join(str(valor).split())

    @staticmethod
    def _split_documento(documento: object) -> Tuple[Optional[str], Optional[str]]:
        """'NFE.175118' -> ('NFE', '175118'); 'CT.CT R205' -> ('CT', 'CT R205')."""
        if documento is None:
            return None, None
        s = " ".join(str(documento).split())
        if not s:
            return None, None
        if "." in s:
            tipo, numero = s.split(".", 1)
            return tipo.strip() or None, numero.strip() or None
        return None, s

    @staticmethod
    def _split_tit_parc(tit_parc: object) -> Tuple[Optional[str], Optional[str]]:
        """'8674/3' -> ('8674', '3'); '8674' -> ('8674', None)."""
        if tit_parc is None:
            return None, None
        s = " ".join(str(tit_parc).split())
        if not s:
            return None, None
        if "/" in s:
            titulo, parcela = s.split("/", 1)
            return titulo.strip() or None, parcela.strip() or None
        return s, None

    @staticmethod
    def _parse_data(valor: object) -> Optional[date]:
        if valor is None:
            return None
        if isinstance(valor, datetime):
            return valor.date()
        if isinstance(valor, date):
            return valor
        s = str(valor).strip()
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})", s)
        if m:
            d, mth, y = (int(g) for g in m.groups())
            try:
                return date(y, mth, d)
            except ValueError:
                return None
        return None

    @staticmethod
    def _is_total_ou_cabecalho(textos: List[str]) -> bool:
        """Detecta linhas de total e cabeçalhos repetidos a partir do texto da linha."""
        joined = " ".join(_normalizar(t) for t in textos if t is not None)
        if not joined.strip():
            return True  # linha em branco
        if "total" in joined:  # 'total do dia', 'total do periodo', 'total geral'
            return True
        if "disponivel em" in joined:
            return True
        # cabeçalho repetido a cada página
        if "documento" in joined and ("tit/parc" in joined or "tit parc" in joined):
            return True
        return False

    def _montar_titulo(
        self,
        data_linha: Optional[date],
        documento: object,
        tit_parc: object,
        orig: object,
        fornecedor: object,
        saidas: float,
    ) -> Titulo:
        tipo_doc, num_doc = self._split_documento(documento)
        numero_titulo, parcela = self._split_tit_parc(tit_parc)
        origem = (str(orig).strip().upper() if orig else None)
        if origem not in ORIGENS_VALIDAS:
            origem = origem or None
        return Titulo(
            id=None,  # vínculo real (ID interno Sienge) é resolvido na etapa de orquestração
            numero=numero_titulo or (str(num_doc) if num_doc else ""),
            fornecedor_nome=self._limpar_texto(fornecedor),
            fornecedor_cnpj="",  # não consta no relatório; vem da NF/Sefaz/API
            valor_nominal=saidas,
            valor_liquido=saidas,
            data_vencimento=data_linha,
            forma_pagamento="",
            status="",
            documento=(self._limpar_texto(documento) or None),
            tipo_documento=tipo_doc,
            numero_documento=num_doc,
            parcela=parcela,
            origem=origem,
            data_referencia=data_linha,
        )

    # ---------------------------------------------------------------- XLSX

    def _parse_xlsx(self, path: str) -> List[Titulo]:
        from openpyxl import load_workbook

        wb = load_workbook(path, data_only=True, read_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()

        # 1. Localizar a linha de cabeçalho da tabela
        header_idx, col_map = self._localizar_cabecalho(rows)
        if header_idx is None:
            logger.error("Cabeçalho da tabela não encontrado no .xlsx (Data/Documento/Tit/Parc).")
            return []

        # 2. Ordenar as colunas conhecidas por índice para montar 'spans'
        ordered = sorted(col_map.items(), key=lambda kv: kv[1])  # [(key, idx), ...]
        total_cols = max(len(r) for r in rows) if rows else 0

        def span(key: str) -> Tuple[int, int]:
            idx = col_map[key]
            seguintes = [i for _, i in ordered if i > idx]
            fim = min(seguintes) if seguintes else total_cols
            return idx, fim

        def valor_span(row: List, key: str) -> object:
            if key not in col_map:
                return None
            ini, fim = span(key)
            for cel in row[ini:fim]:
                if cel is not None and str(cel).strip() != "":
                    return cel
            return None

        titulos: List[Titulo] = []
        total_lidas = lidas_validas = ignoradas_entrada = ignoradas_total = 0

        for row in rows[header_idx + 1:]:
            total_lidas += 1
            if self._is_total_ou_cabecalho([str(c) for c in row]):
                ignoradas_total += 1
                continue

            documento = valor_span(row, "documento")
            tit_parc = valor_span(row, "tit_parc")
            if not documento and not tit_parc:
                ignoradas_total += 1
                continue

            entradas = self._parse_valor_br(valor_span(row, "entradas"))
            saidas = self._parse_valor_br(valor_span(row, "saidas"))

            if entradas > 0 and saidas <= 0:
                ignoradas_entrada += 1
                logger.debug(f"Ignorando entrada (crédito): doc={documento} valor={entradas}")
                continue
            if saidas <= 0:
                ignoradas_total += 1
                continue

            data_linha = self._parse_data(valor_span(row, "data"))
            fornecedor = valor_span(row, "fornecedor")
            orig = valor_span(row, "orig")

            titulos.append(
                self._montar_titulo(data_linha, documento, tit_parc, orig, fornecedor, saidas)
            )
            lidas_validas += 1

        logger.success(
            f"Relatório XLSX: {total_lidas} linhas lidas | {lidas_validas} conferíveis | "
            f"{ignoradas_entrada} entradas ignoradas | {ignoradas_total} totais/branco/cabeçalho ignorados."
        )
        return titulos

    def _localizar_cabecalho(self, rows: List[List]) -> Tuple[Optional[int], dict]:
        for idx, row in enumerate(rows):
            normalizados = {j: _normalizar(c) for j, c in enumerate(row) if c is not None}
            valores = set(normalizados.values())
            if "documento" in valores and ("tit/parc" in valores or "tit parc" in valores):
                col_map = {}
                for j, norm in normalizados.items():
                    key = _HEADER_KEYS.get(norm)
                    if key and key not in col_map:
                        col_map[key] = j
                return idx, col_map
        return None, {}

    # ----------------------------------------------------------------- PDF

    def _parse_pdf(self, path: str) -> List[Titulo]:
        import pdfplumber

        textos_paginas: List[str] = []
        tabelas_titulos: List[Titulo] = []
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                # Preferir extração por tabela
                try:
                    tabelas = page.extract_tables() or []
                except Exception as e:  # noqa: BLE001
                    logger.warning(f"Falha extract_tables na página: {e}")
                    tabelas = []
                for tab in tabelas:
                    tabelas_titulos.extend(self._parse_tabela_pdf(tab))

                texto = page.extract_text() or ""
                if texto:
                    textos_paginas.append(texto)

        if tabelas_titulos:
            logger.success(f"Relatório PDF (tabela): {len(tabelas_titulos)} títulos conferíveis.")
            return tabelas_titulos

        # Fallback: parsing por linhas/regex sobre o texto cru
        logger.info("PDF sem tabelas detectáveis; usando fallback de texto.")
        return self._parse_pdf_text("\n".join(textos_paginas))

    def _parse_tabela_pdf(self, tabela: List[List]) -> List[Titulo]:
        header_idx, col_map = self._localizar_cabecalho(tabela)
        if header_idx is None:
            return []
        ordered = sorted(col_map.items(), key=lambda kv: kv[1])
        total_cols = max(len(r) for r in tabela) if tabela else 0

        def valor(row, key):
            if key not in col_map:
                return None
            idx = col_map[key]
            seguintes = [i for _, i in ordered if i > idx]
            fim = min(seguintes) if seguintes else total_cols
            for cel in row[idx:fim]:
                if cel is not None and str(cel).strip() != "":
                    return cel
            return None

        titulos = []
        for row in tabela[header_idx + 1:]:
            if self._is_total_ou_cabecalho([str(c) for c in row]):
                continue
            documento = valor(row, "documento")
            tit_parc = valor(row, "tit_parc")
            if not documento and not tit_parc:
                continue
            entradas = self._parse_valor_br(valor(row, "entradas"))
            saidas = self._parse_valor_br(valor(row, "saidas"))
            if (entradas > 0 and saidas <= 0) or saidas <= 0:
                continue
            titulos.append(
                self._montar_titulo(
                    self._parse_data(valor(row, "data")),
                    documento, tit_parc, valor(row, "orig"),
                    valor(row, "fornecedor"), saidas,
                )
            )
        return titulos

    def _parse_pdf_text(self, texto: str) -> List[Titulo]:
        """
        Fallback baseado em texto. Cada linha lógica começa com uma data dd/mm/aaaa;
        o nome do fornecedor pode quebrar em múltiplas linhas físicas e é reagrupado.

        A distinção Entrada x Saída é feita pelo *saldo corrente*: a última coluna de
        cada linha é o saldo; se ele diminui em relação à linha anterior, o movimento é
        uma saída (pagamento) e o título é conferível.
        """
        linhas = [l.rstrip() for l in texto.splitlines()]

        # Saldo inicial ("Disponível em DD/MM/AAAA  1.234,56")
        saldo_anterior: Optional[float] = None
        for l in linhas:
            if "disponivel em" in _normalizar(l):
                nums = re.findall(r"-?[\d.]+,\d{2}", l)
                if nums:
                    saldo_anterior = self._parse_valor_br(nums[-1])
                break

        re_data = re.compile(r"^\s*(\d{2}/\d{2}/\d{4})\b")
        re_num = re.compile(r"-?[\d.]+,\d{2}")

        # Reagrupar: juntar linhas de continuação (sem data) à linha anterior.
        # Linhas de total/cabeçalho/"Disponível em" nunca são tratadas como continuação.
        logicas: List[str] = []
        for l in linhas:
            if not l.strip():
                continue
            if re_data.match(l):
                logicas.append(l)
            elif logicas and not self._is_total_ou_cabecalho([l]):
                logicas[-1] += " " + l.strip()

        titulos: List[Titulo] = []
        total = validas = ignoradas = 0
        for linha in logicas:
            if self._is_total_ou_cabecalho([linha]):
                continue
            total += 1
            m = re_data.match(linha)
            data_linha = self._parse_data(m.group(1))
            resto = linha[m.end():].strip()

            tokens = resto.split()
            if len(tokens) < 3:
                ignoradas += 1
                continue
            documento = tokens[0]
            tit_parc = tokens[1]
            orig = tokens[2] if tokens[2].upper() in ORIGENS_VALIDAS else None
            corpo_ini = 3 if orig else 2

            numeros = re_num.findall(resto)
            if not numeros:
                ignoradas += 1
                continue

            # remove os números do corpo para isolar o fornecedor
            corpo = resto
            for n in numeros:
                corpo = corpo.replace(n, " ")
            corpo_tokens = corpo.split()[corpo_ini:]
            fornecedor = " ".join(corpo_tokens).strip()

            saldo_atual = self._parse_valor_br(numeros[-1])
            movimento = self._parse_valor_br(numeros[-2]) if len(numeros) >= 2 else saldo_atual

            eh_saida = False
            if len(numeros) >= 2 and saldo_anterior is not None:
                eh_saida = saldo_atual < saldo_anterior - 1e-6
            elif len(numeros) >= 2:
                eh_saida = True  # sem saldo de referência: assume movimento como saída
            saldo_anterior = saldo_atual if len(numeros) >= 2 else saldo_anterior

            if not eh_saida or movimento <= 0:
                ignoradas += 1
                continue

            titulos.append(
                self._montar_titulo(data_linha, documento, tit_parc, orig, fornecedor, movimento)
            )
            validas += 1

        logger.success(
            f"Relatório PDF (texto): {total} linhas lógicas | {validas} conferíveis | {ignoradas} ignoradas."
        )
        return titulos
